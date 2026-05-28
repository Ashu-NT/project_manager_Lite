from __future__ import annotations

import csv
import os


def export_to_file(
    rows: list[dict[str, object]],
    columns: list[dict[str, str]],
    file_path: str,
) -> dict[str, object]:
    """Export rows to .xlsx (openpyxl) or .csv. Returns {ok, message, path} or {ok, error}."""
    if not rows:
        return {"ok": False, "error": "No records to export."}
    if not columns:
        return {"ok": False, "error": "No columns selected for export."}
    resolved = (file_path or "").strip()
    if not resolved:
        return {"ok": False, "error": "No file path specified."}
    _, ext = os.path.splitext(resolved.lower())
    try:
        if ext == ".xlsx":
            return _write_xlsx(rows, columns, resolved)
        return _write_csv(rows, columns, resolved)
    except Exception as exc:
        return {"ok": False, "error": str(exc)}


def _write_xlsx(
    rows: list[dict[str, object]],
    columns: list[dict[str, str]],
    file_path: str,
) -> dict[str, object]:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D5F8A", end_color="2D5F8A", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col.get("label") or col.get("key", ""))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            key = col.get("key", "")
            val = row.get(key, "")
            if val is None:
                val = ""
            elif not isinstance(val, (int, float, bool)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx, col in enumerate(columns, start=1):
        key = col.get("key", "")
        label = col.get("label") or key
        max_len = max(
            len(label),
            max((len(str(r.get(key) or "")) for r in rows), default=0),
        )
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = min(max_len + 4, 52)

    ws.freeze_panes = "A2"
    wb.save(file_path)
    return {
        "ok": True,
        "message": f"Exported {len(rows)} record(s) to {os.path.basename(file_path)}.",
        "path": file_path,
    }


def _write_csv(
    rows: list[dict[str, object]],
    columns: list[dict[str, str]],
    file_path: str,
) -> dict[str, object]:
    keys = [col.get("key", "") for col in columns]
    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([col.get("label") or col.get("key", "") for col in columns])
        for row in rows:
            writer.writerow([str(row.get(k, "") or "") for k in keys])
    return {
        "ok": True,
        "message": f"Exported {len(rows)} record(s) to {os.path.basename(file_path)}.",
        "path": file_path,
    }


__all__ = ["export_to_file"]
