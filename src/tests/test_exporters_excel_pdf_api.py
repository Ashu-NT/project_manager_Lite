from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.core.platform.domain.security.auth.session import UserSessionPrincipal
from src.core.platform.common.exceptions import BusinessRuleError
from src.core.modules.project_management.domain.enums import DependencyType
from src.core.modules.project_management.infrastructure.reporting import api as reporting_api
from src.core.modules.project_management.infrastructure.reporting.models.contexts import (
    FinanceLedgerExportPage,
    MAX_FINANCE_LEDGER_EXPORT_ROWS,
)
from src.core.modules.project_management.infrastructure.reporting.models import (
    CostSourceBreakdown,
    CostSourceRow,
    CostBreakdownRow,
    EarnedValueMetrics,
    EvmSeriesPoint,
    GanttTaskBar,
    ProjectKPI,
    ResourceLoadRow,
    TaskVarianceRow,
)


def _setup_report_project(services):
    ps = services["project_service"]
    ts = services["task_service"]
    rs = services["resource_service"]
    bs = services["baseline_service"]

    project = ps.create_project(
        "Export Validation Project",
        "",
        start_date=date(2023, 11, 6),
        end_date=date(2023, 11, 30),
        financial_currency_code="USD",
    )
    pid = project.id

    t1 = ts.create_task(pid, "Task Alpha", start_date=date(2023, 11, 6), duration_days=2)
    t2 = ts.create_task(pid, "Task Beta", duration_days=2)
    ts.add_dependency(t1.id, t2.id, DependencyType.FINISH_TO_START, lag_days=0)

    res = rs.create_resource(
        "Exporter Dev",
        "Developer",
        hourly_rate=100.0,
        currency_code="USD",
        rate_effective_on=date(2023, 11, 6),
    )
    assignment = ts.assign_resource(t1.id, res.id, allocation_percent=50.0)
    ts.set_assignment_hours(assignment.id, 4.0)

    baseline = bs.create_baseline(pid, "Baseline Export", rate_as_of=date.today())
    ts.update_progress(t1.id, percent_complete=50.0)
    return pid, baseline.id


def test_excel_export_contains_expected_sections_when_baseline_exists(services, tmp_path):
    pid, baseline_id = _setup_report_project(services)
    output = tmp_path / "report.xlsx"

    reporting_api.generate_excel_report(
        services["reporting_service"],
        pid,
        output,
        baseline_id=baseline_id,
        as_of=date(2023, 11, 30),
    )

    wb = load_workbook(output)
    names = set(wb.sheetnames)
    assert {"Overview", "Tasks", "Resources", "EVM", "Variance", "Cost Sources"}.issubset(names)
    assert "Cost Breakdown" not in names
    assert wb["Overview"]["A1"].value.startswith("Project KPIs - ")
    assert wb["Tasks"]["A1"].value == "Task ID"
    assert wb["EVM"]["A2"].value == "Metric"
    assert wb["EVM"]["D2"].value == "Period End"
    assert wb["Cost Sources"]["A1"].value == "Source"


def test_excel_export_without_baseline_skips_evm_sheet(services, tmp_path):
    ps = services["project_service"]
    ts = services["task_service"]

    project = ps.create_project("No Baseline Export", "")
    pid = project.id
    ts.create_task(pid, "Task No Base", start_date=date(2023, 11, 6), duration_days=2)

    output = tmp_path / "report_no_baseline.xlsx"
    reporting_api.generate_excel_report(services["reporting_service"], pid, output)

    wb = load_workbook(output)
    assert "EVM" not in wb.sheetnames
    assert {"Overview", "Tasks", "Resources"}.issubset(set(wb.sheetnames))


def test_pdf_export_succeeds_when_gantt_generation_fails(services, tmp_path, monkeypatch):
    pid, _baseline_id = _setup_report_project(services)
    temp_dir = tmp_path / "tmp_reports"

    def _raise_gantt(*_args, **_kwargs):
        raise ValueError("No tasks with dates available for Gantt chart.")

    monkeypatch.setattr("src.core.modules.project_management.infrastructure.reporting.api.generate_gantt_png", _raise_gantt)

    output = tmp_path / "report.pdf"
    reporting_api.generate_pdf_report(
        services["reporting_service"],
        pid,
        output,
        temp_dir=temp_dir,
        as_of=date(2023, 11, 30),
    )

    assert output.exists()
    assert output.read_bytes().startswith(b"%PDF")
    assert not temp_dir.exists()


def test_pdf_export_cleans_temporary_gantt_artifact(services, tmp_path):
    pid, _baseline_id = _setup_report_project(services)
    temp_dir = tmp_path / "tmp_reports_cleanup"
    output = tmp_path / "report_cleanup.pdf"

    reporting_api.generate_pdf_report(
        services["reporting_service"],
        pid,
        output,
        temp_dir=temp_dir,
        as_of=date(2023, 11, 30),
    )

    assert output.exists()
    assert output.read_bytes().startswith(b"%PDF")
    assert not (temp_dir / f"gantt_{pid}.png").exists()
    assert not temp_dir.exists()


def test_reporting_api_populates_optional_contexts(monkeypatch, tmp_path):
    kpi = ProjectKPI(
        project_id="p1",
        name="API Context",
        start_date=date(2023, 11, 6),
        end_date=date(2023, 11, 30),
        duration_working_days=19,
        tasks_total=1,
        tasks_completed=0,
        tasks_in_progress=1,
        task_blocked=0,
        tasks_not_started=0,
        critical_tasks=1,
        late_tasks=0,
        total_planned_cost=100.0,
        total_actual_cost=50.0,
        cost_variance=-50.0,
        total_committed_cost=60.0,
        committment_variance=-40.0,
    )
    bars = [
        GanttTaskBar(
            task_id="t1",
            name="Task API",
            start=date(2023, 11, 6),
            end=date(2023, 11, 7),
            is_critical=True,
            percent_complete=50.0,
            status="IN_PROGRESS",
        )
    ]
    resources = [ResourceLoadRow(resource_id="r1", resource_name="Res", total_allocation_percent=50.0, tasks_count=1)]
    evm = EarnedValueMetrics(
        as_of=date(2023, 11, 30),
        baseline_id="b1",
        BAC=100.0,
        PV=80.0,
        EV=60.0,
        AC=50.0,
        CPI=1.2,
        SPI=0.75,
        EAC=83.33,
        ETC=33.33,
        VAC=16.67,
    )
    series = [
        EvmSeriesPoint(
            period_end=date(2023, 11, 30),
            PV=80.0,
            EV=60.0,
            AC=50.0,
            BAC=100.0,
            CPI=1.2,
            SPI=0.75,
        )
    ]
    variance = [
        TaskVarianceRow(
            task_id="t1",
            task_name="Task API",
            baseline_start=date(2023, 11, 6),
            baseline_finish=date(2023, 11, 7),
            current_start=date(2023, 11, 6),
            current_finish=date(2023, 11, 8),
            start_variance_days=0,
            finish_variance_days=1,
            is_critical=True,
        )
    ]
    cost_breakdown = [
        CostBreakdownRow(cost_type="MATERIAL", currency="USD", planned=100.0, actual=50.0)
    ]
    cost_sources = CostSourceBreakdown(
        project_id="p1",
        project_currency="USD",
        rows=[
            CostSourceRow(
                source_key="DIRECT_COST",
                source_label="Direct Cost",
                planned=100.0,
                committed=60.0,
                actual=50.0,
                forecast=0.0,
            ),
            CostSourceRow(
                source_key="COMPUTED_LABOR",
                source_label="Computed Labor",
                planned=20.0,
                committed=0.0,
                actual=10.0,
                forecast=0.0,
            ),
            CostSourceRow(
                source_key="LABOR_ADJUSTMENT",
                source_label="Labor Adjustment",
                planned=0.0,
                committed=0.0,
                actual=0.0,
                forecast=0.0,
            ),
        ],
        total_planned=120.0,
        total_committed=60.0,
        total_actual=60.0,
        notes=[],
    )

    class DummyReportingService:
        def get_project_kpis(self, _project_id):
            return kpi

        def get_gantt_data(self, _project_id):
            return bars

        def get_resource_load_summary(self, _project_id):
            return resources

        def get_earned_value(self, _project_id, baseline_id=None, as_of=None):
            assert baseline_id == "b1"
            assert as_of == date(2023, 11, 30)
            return evm

        def get_evm_series(self, _project_id, baseline_id=None, as_of=None):
            assert baseline_id == "b1"
            assert as_of == date(2023, 11, 30)
            return series

        def get_baseline_schedule_variance(self, _project_id, baseline_id=None):
            assert baseline_id == "b1"
            return variance

        def get_cost_breakdown(self, _project_id, as_of=None, baseline_id=None):
            assert baseline_id == "b1"
            assert as_of == date(2023, 11, 30)
            return cost_breakdown

        def get_project_cost_source_breakdown(self, _project_id, as_of=None):
            assert as_of == date(2023, 11, 30)
            return cost_sources

    captured = {}

    class _FakeExcelRenderer:
        def render(self, ctx, output_path):
            captured["excel_ctx"] = ctx
            captured["excel_path"] = output_path
            return output_path

    class _FakePdfRenderer:
        def render(self, ctx, output_path):
            captured["pdf_ctx"] = ctx
            captured["pdf_path"] = output_path
            return output_path

    monkeypatch.setattr(reporting_api, "ExcelReportRenderer", lambda: _FakeExcelRenderer())
    monkeypatch.setattr(reporting_api, "PdfReportRenderer", lambda: _FakePdfRenderer())
    monkeypatch.setattr(reporting_api, "generate_gantt_png", lambda *_args, **_kwargs: Path(tmp_path / "gantt.png"))

    service = DummyReportingService()
    as_of = date(2023, 11, 30)

    reporting_api.generate_excel_report(service, "p1", tmp_path / "api.xlsx", baseline_id="b1", as_of=as_of)
    reporting_api.generate_pdf_report(service, "p1", tmp_path / "api.pdf", baseline_id="b1", as_of=as_of)

    assert captured["excel_ctx"].evm is evm
    assert captured["excel_ctx"].evm_series == series
    assert captured["excel_ctx"].baseline_variance == variance
    assert captured["excel_ctx"].cost_breakdown == cost_breakdown
    assert captured["excel_ctx"].cost_sources == cost_sources
    assert captured["excel_ctx"].as_of == as_of

    assert captured["pdf_ctx"].evm is evm
    assert captured["pdf_ctx"].evm_series == series
    assert captured["pdf_ctx"].baseline_variance == variance
    assert captured["pdf_ctx"].cost_breakdown == cost_breakdown
    assert captured["pdf_ctx"].cost_sources == cost_sources
    assert captured["pdf_ctx"].as_of == as_of


def test_reporting_api_requires_report_export_permission_from_live_session(services, tmp_path):
    tenant_id = services["user_session"].stored_active_tenant_id()
    organization_id = services["user_session"].stored_active_organization_id()
    services["module_catalog_service"].set_module_state(
        "project_management",
        licensed=True,
        enabled=True,
    )
    services["user_session"].set_principal(
        UserSessionPrincipal(
            user_id="u-report",
            username="report-viewer",
            display_name="Report Viewer",
            role_names=frozenset({"viewer"}),
            permissions=frozenset({"report.view"}),
            active_tenant_id=tenant_id,
            active_organization_id=organization_id,
        )
    )

    with pytest.raises(BusinessRuleError, match="Permission denied") as exc:
        reporting_api.generate_gantt_png(
            services["reporting_service"],
            "project-1",
            tmp_path / "restricted.png",
        )

    assert exc.value.code == "PERMISSION_DENIED"


def test_excel_export_omits_finance_sections_without_finance_read(services, tmp_path):
    """F0: report.export alone (no finance.read/finance.export) must still
    produce a valid export — the Project Finance sections are simply
    omitted rather than the whole export failing."""
    pid, baseline_id = _setup_report_project(services)
    tenant_id = services["user_session"].stored_active_tenant_id()
    organization_id = services["user_session"].stored_active_organization_id()
    services["user_session"].set_principal(
        UserSessionPrincipal(
            user_id="export-no-finance",
            username="export-no-finance",
            display_name="Export No Finance",
            role_names=frozenset({"viewer"}),
            permissions=frozenset({"report.view", "report.export"}),
            active_tenant_id=tenant_id,
            active_organization_id=organization_id,
        )
    )

    output = tmp_path / "restricted_report.xlsx"
    reporting_api.generate_excel_report(
        services["reporting_service"],
        pid,
        output,
        finance_service=services["finance_service"],
        baseline_id=baseline_id,
        as_of=date(2023, 11, 30),
    )

    wb = load_workbook(output)
    names = set(wb.sheetnames)
    assert {"Overview", "Tasks", "Resources"}.issubset(names)
    assert "EVM" not in names
    assert "Cost Sources" not in names
    assert "Cost Breakdown" not in names
    assert "Finance" not in names
    assert "Finance Ledger" not in names

    overview = wb["Overview"]
    planned_cost_row = next(
        row
        for row in overview.iter_rows(min_col=1, max_col=1)
        if row[0].value == "Planned cost"
    )
    restricted_value = overview.cell(row=planned_cost_row[0].row, column=2).value
    assert restricted_value == "Restricted (finance.read required)"


def test_finance_ledger_export_page_rejects_unbounded_requests():
    with pytest.raises(ValueError, match="non-negative"):
        FinanceLedgerExportPage.build([], offset=-1, limit=1)

    with pytest.raises(ValueError, match="between 1"):
        FinanceLedgerExportPage.build(
            [],
            offset=0,
            limit=MAX_FINANCE_LEDGER_EXPORT_ROWS + 1,
        )
