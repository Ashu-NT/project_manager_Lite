from __future__ import annotations

from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from src.core.modules.project_management.infrastructure.reporting import api as reporting_api
from src.core.modules.project_management.domain.financials.forecast import (
    ForecastGenerationMode,
    ForecastLineSourceKind,
    ForecastLineSourceType,
)


def _approved_controls(services):
    organization = services["tenant_context_service"].get_active_organization()
    project = services["project_service"].create_project(
        "Canonical finance read model",
        financial_currency_code=organization.base_currency,
    )
    code = services["financial_configuration_service"].create_cost_code(
        code="D4-CONTROL",
        name="D.4 control",
    )

    budgets = services["budget_service"]
    budget = budgets.create_budget(project.id, "Approved budget")
    budgets.add_line(
        budget.id,
        cost_code_id=code.id,
        description="Authorized scope",
        amount=Decimal("100"),
        expected_budget_version=budget.row_version,
    )
    budget = budgets.get_budget(budget.id)
    budget = budgets.submit_budget(
        budget.id, "admin", expected_version=budget.row_version
    )
    result = budgets.approve_budget(
        budget.id, approved_by="admin", expected_version=budget.row_version
    )
    budget = budgets.get_budget(result.budget_id)

    forecasts = services["forecast_version_service"]
    forecast = forecasts.create_forecast(
        project.id,
        name="Approved ETC",
        as_of_date=date(2026, 8, 1),
        generation_mode=ForecastGenerationMode.MANUAL,
        created_by="admin",
    )
    forecasts.add_line(
        forecast.id,
        cost_code_id=code.id,
        description="Remaining delivery",
        amount=Decimal("80"),
        source_kind=ForecastLineSourceKind.MANUAL,
        source_type=ForecastLineSourceType.MANUAL_ESTIMATE,
        created_by="admin",
        expected_forecast_version=forecast.row_version,
        period_start=date(2026, 8, 1),
        period_end=date(2026, 8, 31),
    )
    forecast = forecasts.get_forecast(forecast.id)
    forecast = forecasts.submit_forecast(
        forecast.id,
        submitted_by="admin",
        expected_version=forecast.row_version,
    )
    forecast = forecasts.approve_forecast(
        forecast.id,
        approved_by="admin",
        expected_version=forecast.row_version,
    )

    services["financial_period_service"].create_period(
        code="D4-2026-08",
        name="August 2026",
        fiscal_year=2026,
        period_number=8,
        start_date=date(2026, 8, 1),
        end_date=date(2026, 8, 31),
    )
    entries = services["cost_entry_service"]
    entry = entries.create_manual_entry(
        project_id=project.id,
        command_id="d4-actual-1",
        description="Posted actual",
        amount=Decimal("25"),
        currency_code=organization.base_currency,
        transaction_date=date(2026, 8, 5),
        cost_code_id=code.id,
    )
    entry = entries.submit(entry.id, expected_version=entry.row_version)
    entries.approve(entry.id, expected_version=entry.row_version)
    entry = entries.get_entry(entry.id)
    entry = entries.post(
        entry.id,
        expected_version=entry.row_version,
        posting_date=date(2026, 8, 5),
    )
    return project, budget, forecast, entry, code


def test_snapshot_reconciles_approved_budget_forecast_and_posted_actual(services) -> None:
    project, budget, forecast, _entry, _code = _approved_controls(services)

    snapshot = services["finance_service"].get_finance_snapshot(
        project.id, as_of=date(2026, 8, 31)
    )

    assert snapshot.budget == Decimal("100")
    assert snapshot.actual == Decimal("25")
    assert snapshot.forecast_etc == Decimal("80")
    assert snapshot.estimate_at_completion == Decimal("105")
    assert snapshot.variance_at_completion == Decimal("-5")
    assert snapshot.approved_budget_id == budget.id
    assert snapshot.approved_forecast_id == forecast.id
    assert snapshot.approved_forecast_revision == forecast.revision
    assert snapshot.currency_basis == "PROJECT_CURRENCY"
    assert snapshot.period_granularity == "month"
    assert snapshot.reconciliation.is_reconciled is True
    assert snapshot.reconciliation.posted_actual_delta == Decimal("0")
    assert snapshot.reconciliation.open_commitment_delta == Decimal("0")
    assert snapshot.reconciliation.forecast_etc_delta == Decimal("0")
    assert sum(
        (row.amount for row in snapshot.ledger if row.stage == "forecast"),
        start=Decimal("0"),
    ) == snapshot.forecast_etc
    august = next(row for row in snapshot.cost_phasing if row.period_key == "2026-08")
    assert august.actual == Decimal("25")
    assert august.forecast == Decimal("80")
    assert august.exposure == Decimal("105")


def test_finance_excel_export_has_bounded_lineage_and_control_parity(
    services,
    tmp_path,
) -> None:
    project, budget, forecast, _entry, code = _approved_controls(services)
    output = tmp_path / "finance-d5.xlsx"

    reporting_api.generate_excel_report(
        services["reporting_service"],
        project.id,
        output,
        finance_service=services["finance_service"],
        as_of=date(2026, 8, 31),
        finance_ledger_offset=1,
        finance_ledger_limit=1,
    )

    workbook = load_workbook(output, data_only=True)
    finance = workbook["Finance"]
    metadata = {
        finance.cell(row=row, column=1).value: finance.cell(row=row, column=2).value
        for row in range(1, finance.max_row + 1)
    }
    assert metadata["Snapshot as of"] == "2026-08-31"
    assert metadata["Currency basis"].startswith("PROJECT_CURRENCY:")
    assert metadata["Approved budget version"] == f"{budget.id} / revision {budget.revision}"
    assert metadata["Approved forecast version"] == f"{forecast.id} / revision {forecast.revision}"
    assert metadata["Reconciliation status"] == "Reconciled"
    assert metadata["Ledger page limit"] == 1

    ledger = workbook["Finance Ledger"]
    headers = [cell.value for cell in ledger[1]]
    assert headers == [
        "Date",
        "Period Start",
        "Period End",
        "Source",
        "Source Type",
        "Stage",
        "Cost Type",
        "Cost Code ID",
        "Financial Period ID",
        "Reference Type",
        "Reference ID",
        "Reference",
        "Task ID",
        "Task",
        "Resource ID",
        "Resource",
        "Amount",
        "Currency",
    ]
    assert ledger.max_row == 2
    assert ledger.cell(row=2, column=8).value == code.id
    assert ledger.cell(row=2, column=9).value


def test_finance_pdf_export_uses_the_same_canonical_read_basis(
    services,
    tmp_path,
) -> None:
    project, _budget, _forecast, _entry, _code = _approved_controls(services)
    output = tmp_path / "finance-d5.pdf"

    reporting_api.generate_pdf_report(
        services["reporting_service"],
        project.id,
        output,
        temp_dir=tmp_path / "report-temp",
        finance_service=services["finance_service"],
        as_of=date(2026, 8, 31),
        finance_ledger_limit=1,
    )

    assert output.read_bytes().startswith(b"%PDF")


def test_snapshot_has_no_eac_or_vac_before_the_approved_forecast_basis(services) -> None:
    project, _budget, _forecast, _entry, _code = _approved_controls(services)

    snapshot = services["finance_service"].get_finance_snapshot(
        project.id, as_of=date(2026, 7, 31)
    )

    assert snapshot.actual == Decimal("0")
    assert snapshot.forecast_etc is None
    assert snapshot.estimate_at_completion is None
    assert snapshot.variance_at_completion is None
    assert snapshot.approved_forecast_id is None


def test_posted_reversal_nets_actual_without_rewriting_forecast(services) -> None:
    project, _budget, forecast, entry, _code = _approved_controls(services)
    entries = services["cost_entry_service"]
    entries.reverse(
        entry.id,
        expected_version=entry.row_version,
        command_id="d4-reversal-1",
        posting_date=date(2026, 8, 20),
        reason="Correct the test posting",
    )

    snapshot = services["finance_service"].get_finance_snapshot(
        project.id, as_of=date(2026, 8, 31)
    )

    assert snapshot.actual == Decimal("0")
    assert snapshot.forecast_etc == Decimal("80")
    assert snapshot.estimate_at_completion == Decimal("80")
    assert snapshot.approved_forecast_id == forecast.id


def test_as_of_selects_superseded_approved_forecast_version(services) -> None:
    project, _budget, first, _entry, code = _approved_controls(services)
    forecasts = services["forecast_version_service"]
    second = forecasts.create_forecast(
        project.id,
        name="September ETC",
        as_of_date=date(2026, 9, 1),
        generation_mode=ForecastGenerationMode.MANUAL,
        created_by="admin",
    )
    forecasts.add_line(
        second.id,
        cost_code_id=code.id,
        description="Revised remaining delivery",
        amount=Decimal("60"),
        source_kind=ForecastLineSourceKind.MANUAL,
        source_type=ForecastLineSourceType.MANUAL_ESTIMATE,
        created_by="admin",
        expected_forecast_version=second.row_version,
    )
    second = forecasts.get_forecast(second.id)
    second = forecasts.submit_forecast(
        second.id, submitted_by="admin", expected_version=second.row_version
    )
    second = forecasts.approve_forecast(
        second.id, approved_by="admin", expected_version=second.row_version
    )

    august = services["finance_service"].get_finance_snapshot(
        project.id, as_of=date(2026, 8, 31)
    )
    september = services["finance_service"].get_finance_snapshot(
        project.id, as_of=date(2026, 9, 30)
    )

    assert august.approved_forecast_id == first.id
    assert august.forecast_etc == Decimal("80")
    assert september.approved_forecast_id == second.id
    assert september.forecast_etc == Decimal("60")
