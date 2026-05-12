from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from src.core.platform.report_runtime import ChartBlock, MetricBlock, ReportDocument, TableBlock, TextBlock


def render_report_document_excel(document: ReportDocument, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill("solid", fgColor="DDE7F5")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    row = 1
    summary.cell(row=row, column=1, value=document.title).font = title_font
    row += 2

    table_sheets: set[str] = set()
    for section in document.sections:
        summary.cell(row=row, column=1, value=section.title).font = header_font
        row += 1
        for block in section.blocks:
            if isinstance(block, MetricBlock):
                summary.cell(row=row, column=1, value=block.title).font = header_font
                row += 1
                for metric in block.rows:
                    summary.cell(row=row, column=1, value=metric.label).border = thin
                    summary.cell(row=row, column=2, value=str(metric.value)).border = thin
                    row += 1
                row += 1
            elif isinstance(block, TextBlock):
                summary.cell(row=row, column=1, value=block.title).font = header_font
                row += 1
                summary.cell(row=row, column=1, value=block.body).alignment = Alignment(wrap_text=True, vertical="top")
                row += 2
            elif isinstance(block, ChartBlock):
                summary.cell(row=row, column=1, value=block.title).font = header_font
                row += 1
                summary.cell(row=row, column=1, value=f"{block.chart_type.title()} chart defined in report model").alignment = Alignment(wrap_text=True)
                row += 2
            elif isinstance(block, TableBlock):
                sheet_name = _unique_sheet_name(table_sheets, block.title)
                sheet = workbook.create_sheet(sheet_name)
                sheet.cell(row=1, column=1, value=block.title).font = title_font
                for index, header in enumerate(block.columns, start=1):
                    cell = sheet.cell(row=3, column=index, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin
                for data_row_index, data_row in enumerate(block.rows, start=4):
                    for column_index, value in enumerate(data_row, start=1):
                        cell = sheet.cell(row=data_row_index, column=column_index, value=_excel_value(value))
                        cell.border = thin
                        if isinstance(value, (int, float)):
                            cell.alignment = Alignment(horizontal="right")
                for column_index in range(1, len(block.columns) + 1):
                    sheet.column_dimensions[_column_letter(column_index)].width = 22

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 18
    workbook.save(output_path)
    return output_path


def render_report_document_pdf(document: ReportDocument, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=36,
        rightMargin=36,
        topMargin=32,
        bottomMargin=32,
    )
    styles = getSampleStyleSheet()
    story = [Paragraph(document.title, styles["Title"]), Spacer(1, 12)]

    for section in document.sections:
        story.append(Paragraph(section.title, styles["Heading2"]))
        story.append(Spacer(1, 6))
        for block in section.blocks:
            if isinstance(block, MetricBlock):
                story.append(Paragraph(block.title, styles["Heading3"]))
                rows = [["Metric", "Value"]]
                rows.extend([[metric.label, str(metric.value)] for metric in block.rows])
                story.append(_pdf_table(rows))
                story.append(Spacer(1, 10))
            elif isinstance(block, TextBlock):
                story.append(Paragraph(block.title, styles["Heading3"]))
                story.append(Paragraph(block.body, styles["BodyText"]))
                story.append(Spacer(1, 10))
            elif isinstance(block, ChartBlock):
                story.append(Paragraph(block.title, styles["Heading3"]))
                story.append(Paragraph(f"{block.chart_type.title()} chart defined in report model.", styles["BodyText"]))
                story.append(Spacer(1, 10))
            elif isinstance(block, TableBlock):
                story.append(Paragraph(block.title, styles["Heading3"]))
                rows = [list(block.columns)]
                rows.extend([[_pdf_value(value) for value in data_row] for data_row in block.rows])
                story.append(_pdf_table(rows))
                story.append(Spacer(1, 10))

    doc.build(story)
    return output_path


def _pdf_table(rows: list[list[object]]) -> Table:
    table = Table(rows)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    return table


def _excel_value(value: object) -> object:
    if hasattr(value, "isoformat") and value is not None:
        return value.isoformat()
    return value


def _pdf_value(value: object) -> str:
    if hasattr(value, "isoformat") and value is not None:
        return value.isoformat()
    return str(value)


def _unique_sheet_name(existing: set[str], title: str) -> str:
    base = _sanitize_sheet_name(title)
    candidate = base
    index = 2
    while candidate in existing:
        suffix = f" {index}"
        candidate = f"{base[: max(1, 31 - len(suffix))]}{suffix}"
        index += 1
    existing.add(candidate)
    return candidate


def _sanitize_sheet_name(value: str) -> str:
    cleaned = "".join("-" if char in "\\/*?:[]" else char for char in (value or "Report")).strip()
    return (cleaned or "Report")[:31]


def _column_letter(index: int) -> str:
    letters = ""
    value = index
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


__all__ = [
    "render_report_document_excel",
    "render_report_document_pdf",
]
