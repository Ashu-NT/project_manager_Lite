from __future__ import annotations

from pathlib import Path
from uuid import uuid4

from PySide6.QtWidgets import QDialog, QSizePolicy
from openpyxl import load_workbook

from core.platform.party.domain import PartyType
from tests.ui_runtime_helpers import login_as, make_settings_store, register_and_login
from ui.modules.inventory_procurement.data_exchange.import_dialog import InventoryImportDialog
from ui.modules.inventory_procurement.data_exchange_tab import InventoryDataExchangeTab
from ui.modules.inventory_procurement.dashboard_tab import InventoryDashboardTab
from ui.modules.inventory_procurement.item_dialogs import InventoryItemEditDialog
from ui.modules.inventory_procurement.items_tab import InventoryItemsTab
from ui.modules.inventory_procurement.movements_tab import MovementsTab
from ui.modules.inventory_procurement.purchase_orders_tab import PurchaseOrdersTab
from ui.modules.inventory_procurement.receiving_tab import ReceivingTab
from ui.modules.inventory_procurement.reservations_tab import ReservationsTab
from ui.modules.inventory_procurement.requisitions_tab import RequisitionsTab
from ui.modules.inventory_procurement.reports_tab import InventoryReportsTab
from ui.modules.inventory_procurement.stock_tab import StockTab
from ui.modules.inventory_procurement.storeroom_dialogs import StoreroomEditDialog
from ui.platform.shell.main_window import MainWindow
from ui.platform.shared.styles.theme import base_stylesheet


def _child_labels(item) -> list[str]:
    return [item.child(i).text(0) for i in range(item.childCount())]


def _select_combo_value(combo, value) -> None:
    index = combo.findData(value)
    assert index >= 0
    combo.setCurrentIndex(index)


def _mute_message_boxes(monkeypatch) -> None:
    monkeypatch.setattr("PySide6.QtWidgets.QMessageBox.information", lambda *args, **kwargs: None)
    monkeypatch.setattr("PySide6.QtWidgets.QMessageBox.warning", lambda *args, **kwargs: None)
    monkeypatch.setattr("PySide6.QtWidgets.QMessageBox.critical", lambda *args, **kwargs: None)


def _enable_inventory_module(services) -> None:
    services["module_catalog_service"].set_module_state(
        "inventory_procurement",
        licensed=True,
        enabled=True,
    )


def _create_procurement_context(services):
    suffix = uuid4().hex[:6].upper()
    site = services["site_service"].create_site(
        site_code=f"INV-{suffix}",
        name=f"Inventory Site {suffix}",
        currency_code="EUR",
    )
    item = services["inventory_item_service"].create_item(
        item_code=f"MOTOR-{suffix}",
        name=f"Motor {suffix}",
        status="ACTIVE",
        stock_uom="EA",
        is_purchase_allowed=True,
    )
    storeroom = services["inventory_service"].create_storeroom(
        storeroom_code=f"MAIN-{suffix}",
        name=f"Main Stores {suffix}",
        site_id=site.id,
        status="ACTIVE",
    )
    supplier = services["party_service"].create_party(
        party_code=f"SUP-{suffix}",
        party_name=f"Supplier {suffix}",
        party_type=PartyType.SUPPLIER,
    )
    return site, storeroom, item, supplier


def _create_purchase_flow(services, *, with_receipt: bool = False):
    auth = services["auth_service"]
    manager_name = f"inventory-ui-manager-{uuid4().hex[:6]}"
    approver_name = f"inventory-ui-approver-{uuid4().hex[:6]}"
    auth.register_user(manager_name, "StrongPass123", role_names=["inventory_manager"])
    auth.register_user(approver_name, "StrongPass123", role_names=["approver"])
    site, storeroom, item, supplier = _create_procurement_context(services)
    procurement = services["inventory_procurement_service"]
    purchasing = services["inventory_purchasing_service"]
    approvals = services["approval_service"]

    login_as(services, manager_name, "StrongPass123")
    requisition = procurement.create_requisition(
        requesting_site_id=site.id,
        requesting_storeroom_id=storeroom.id,
        purpose="Restock motors",
    )
    requisition_line = procurement.add_requisition_line(
        requisition.id,
        stock_item_id=item.id,
        quantity_requested=5,
        suggested_supplier_party_id=supplier.id,
        estimated_unit_cost=120.0,
    )
    requisition = procurement.submit_requisition(requisition.id)

    login_as(services, approver_name, "StrongPass123")
    approvals.approve_and_apply(requisition.approval_request_id, note="Approved requisition")

    login_as(services, manager_name, "StrongPass123")
    purchase_order = purchasing.create_purchase_order(
        site_id=site.id,
        supplier_party_id=supplier.id,
        currency_code="EUR",
        source_requisition_id=requisition.id,
    )
    purchase_order_line = purchasing.add_purchase_order_line(
        purchase_order.id,
        stock_item_id=item.id,
        destination_storeroom_id=storeroom.id,
        quantity_ordered=5,
        unit_price=115.0,
        source_requisition_line_id=requisition_line.id,
    )
    purchase_order = purchasing.submit_purchase_order(purchase_order.id)

    login_as(services, approver_name, "StrongPass123")
    approvals.approve_and_apply(purchase_order.approval_request_id, note="Approved purchase order")

    receipt = None
    if with_receipt:
        login_as(services, manager_name, "StrongPass123")
        receipt = purchasing.post_receipt(
            purchase_order.id,
            receipt_lines=[
                {
                    "purchase_order_line_id": purchase_order_line.id,
                    "quantity_accepted": 3,
                    "quantity_rejected": 1,
                    "unit_cost": 115.0,
                }
            ],
            supplier_delivery_reference="DEL-UI-01",
        )
    else:
        login_as(services, manager_name, "StrongPass123")
    return site, storeroom, item, supplier, requisition, purchase_order, purchase_order_line, receipt


def test_inventory_items_tab_shows_platform_referenced_supplier_context(qapp, services):
    _enable_inventory_module(services)
    supplier = services["party_service"].create_party(
        party_code="SUP-INV-01",
        party_name="North Supply",
        party_type=PartyType.SUPPLIER,
    )
    register_and_login(services, username_prefix="inventory-ui-items", role_names=("inventory_manager",))
    services["inventory_item_service"].create_item(
        item_code="FILTER-01",
        name="Filter Cartridge",
        status="ACTIVE",
        stock_uom="EA",
        preferred_party_id=supplier.id,
    )

    tab = InventoryItemsTab(
        item_service=services["inventory_item_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.table.rowCount() == 1
    assert tab.count_badge.text() == "1 items"
    assert tab.active_badge.text() == "1 active"
    assert tab.stocked_badge.text() == "1 stocked"
    assert tab.table.item(0, 5).text() == "SUP-INV-01 - North Supply"

    tab.table.selectRow(0)
    qapp.processEvents()

    assert tab.detail_name.text() == "FILTER-01 - Filter Cartridge"
    assert "Preferred party: SUP-INV-01 - North Supply" in tab.detail_status.text()
    assert tab.detail_documents.text() == "No linked documents"


def test_inventory_dashboard_tab_surfaces_stock_pressure_and_procurement_queue(qapp, services):
    _enable_inventory_module(services)
    (
        _site,
        storeroom,
        _item,
        _supplier,
        _requisition,
        purchase_order,
        _purchase_order_line,
        _receipt,
    ) = _create_purchase_flow(services)
    low_stock_item = services["inventory_item_service"].create_item(
        item_code="FILTER-DASH-01",
        name="Filter Cartridge",
        status="ACTIVE",
        stock_uom="EA",
        reorder_point=2,
        reorder_qty=5,
        is_purchase_allowed=True,
    )
    services["inventory_stock_service"].post_opening_balance(
        stock_item_id=low_stock_item.id,
        storeroom_id=storeroom.id,
        quantity=1.0,
        unit_cost=10.0,
    )
    services["inventory_reservation_service"].create_reservation(
        stock_item_id=low_stock_item.id,
        storeroom_id=storeroom.id,
        reserved_qty=1.0,
        source_reference_type="maintenance_task",
        source_reference_id="MT-100",
    )

    tab = InventoryDashboardTab(
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        stock_service=services["inventory_stock_service"],
        reservation_service=services["inventory_reservation_service"],
        procurement_service=services["inventory_procurement_service"],
        purchasing_service=services["inventory_purchasing_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.low_stock_badge.text() == "1 low stock"
    assert tab.awaiting_badge.text() == "0 awaiting approval"
    assert tab.receiving_badge.text() == "1 open receiving"
    assert tab.on_order_badge.text() == "On order: 5.000"
    assert tab.low_stock_table.rowCount() == 1
    assert low_stock_item.item_code in tab.low_stock_table.item(0, 0).text()
    assert tab.queue_table.rowCount() >= 1
    queue_numbers = {tab.queue_table.item(row, 1).text() for row in range(tab.queue_table.rowCount())}
    assert purchase_order.po_number in queue_numbers


def test_inventory_import_dialog_previews_and_imports_storerooms_csv(qapp, services, tmp_path, monkeypatch):
    _enable_inventory_module(services)
    _mute_message_boxes(monkeypatch)
    site = services["site_service"].create_site(site_code="UI-IMP", name="UI Import Site", currency_code="EUR")
    supplier = services["party_service"].create_party(
        party_code="UI-SUP",
        party_name="UI Import Supplier",
        party_type=PartyType.SUPPLIER,
    )
    register_and_login(services, username_prefix="inventory-ui-import", role_names=("inventory_manager",))

    csv_path = tmp_path / "storerooms-ui.csv"
    csv_path.write_text(
        "storeroom_code,name,site_code,status,manager_party_code,allows_receiving\n"
        f"UI-IMP-MAIN,Imported Main,{site.site_code},ACTIVE,{supplier.party_code},true\n",
        encoding="utf-8",
    )

    dialog = InventoryImportDialog(data_exchange_service=services["inventory_data_exchange_service"])
    _select_combo_value(dialog.type_combo, "storerooms")
    dialog.file_path_edit.setText(str(csv_path))
    dialog._load_columns_into_mapping()

    dialog.preview_import()

    assert dialog.preview_table.rowCount() == 1
    assert "1 ready rows" in dialog.summary_label.text()

    dialog.execute_import()

    storerooms = services["inventory_service"].list_storerooms()
    assert any(storeroom.storeroom_code == "UI-IMP-MAIN" for storeroom in storerooms)


def test_inventory_data_exchange_tab_exports_item_and_purchase_order_feeds(qapp, services, tmp_path, monkeypatch):
    _enable_inventory_module(services)
    _mute_message_boxes(monkeypatch)
    (
        _site,
        _storeroom,
        item,
        _supplier,
        _requisition,
        purchase_order,
        _purchase_order_line,
        _receipt,
    ) = _create_purchase_flow(services)

    tab = InventoryDataExchangeTab(
        data_exchange_service=services["inventory_data_exchange_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    _select_combo_value(tab.export_type_combo, "items")
    items_path = tab.export_selected_csv(tmp_path / "inventory-items.csv")
    _select_combo_value(tab.export_type_combo, "purchase_orders")
    purchase_order_path = tab.export_selected_csv(tmp_path / "inventory-purchase-orders.csv")

    assert items_path is not None
    assert purchase_order_path is not None
    assert item.item_code in items_path.read_text(encoding="utf-8")
    assert purchase_order.po_number in purchase_order_path.read_text(encoding="utf-8")


def test_inventory_reports_tab_exports_stock_and_procurement_packages(qapp, services, tmp_path, monkeypatch):
    _enable_inventory_module(services)
    _mute_message_boxes(monkeypatch)
    (
        site,
        storeroom,
        item,
        supplier,
        _requisition,
        _purchase_order,
        _purchase_order_line,
        _receipt,
    ) = _create_purchase_flow(services, with_receipt=True)

    tab = InventoryReportsTab(
        reporting_service=services["inventory_reporting_service"],
        reference_service=services["inventory_reference_service"],
        inventory_service=services["inventory_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    assert tab.report_header_card.sizePolicy().verticalPolicy() == QSizePolicy.Fixed
    assert tab.report_controls_card.sizePolicy().verticalPolicy() == QSizePolicy.Fixed
    _select_combo_value(tab.site_combo, site.id)
    _select_combo_value(tab.storeroom_combo, storeroom.id)
    _select_combo_value(tab.supplier_combo, supplier.id)

    stock_path = tab.export_stock_status_csv(tmp_path / "inventory-stock.csv")
    procurement_path = tab.export_procurement_overview_excel(tmp_path / "inventory-procurement.xlsx")

    assert stock_path is not None
    assert procurement_path is not None
    assert item.item_code in stock_path.read_text(encoding="utf-8")
    workbook = load_workbook(procurement_path)
    assert {"Summary", "Requisitions", "Purchase Orders", "Receipts"} <= set(workbook.sheetnames)


def test_new_item_and_storeroom_dialogs_default_to_active_status(qapp, services):
    site = services["site_service"].create_site(site_code="UID", name="UI Defaults", currency_code="EUR")
    supplier = services["party_service"].create_party(
        party_code="SUP-UID",
        party_name="UI Default Supplier",
        party_type=PartyType.SUPPLIER,
    )

    item_dialog = InventoryItemEditDialog(
        party_options=[("SUP-UID - UI Default Supplier", supplier.id)],
    )
    storeroom_dialog = StoreroomEditDialog(
        site_options=[("UID - UI Defaults", site.id)],
        manager_options=[("SUP-UID - UI Default Supplier", supplier.id)],
    )

    assert item_dialog.status == "ACTIVE"
    assert storeroom_dialog.status == "ACTIVE"


def test_stock_tab_shows_balance_and_transaction_history(qapp, services):
    _enable_inventory_module(services)
    site = services["site_service"].create_site(site_code="HAM", name="Hamburg Warehouse")
    register_and_login(services, username_prefix="inventory-ui-stock", role_names=("inventory_manager",))
    item = services["inventory_item_service"].create_item(
        item_code="BOLT-01",
        name="Hex Bolt",
        status="ACTIVE",
        stock_uom="EA",
    )
    storeroom = services["inventory_service"].create_storeroom(
        storeroom_code="MAIN",
        name="Main Stores",
        site_id=site.id,
        status="ACTIVE",
    )
    services["inventory_stock_service"].post_opening_balance(
        stock_item_id=item.id,
        storeroom_id=storeroom.id,
        quantity=12.0,
        unit_cost=1.25,
    )
    services["inventory_stock_service"].post_adjustment(
        stock_item_id=item.id,
        storeroom_id=storeroom.id,
        quantity=2.0,
        direction="INCREASE",
        unit_cost=1.25,
        reference_id="COUNT-01",
    )

    tab = StockTab(
        stock_service=services["inventory_stock_service"],
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.work_tabs.count() == 2
    assert tab.work_tabs.tabText(0) == "Stock Positions"
    assert tab.work_tabs.tabText(1) == "Transaction History"
    assert tab.balance_table.rowCount() == 1
    assert tab.balance_count_badge.text() == "1 balances"
    assert tab.on_hand_badge.text() == "On hand: 14.000"
    assert tab.available_badge.text() == "Available: 14.000"

    tab.balance_table.selectRow(0)
    qapp.processEvents()

    assert tab.selection_badge.text().startswith("Selection: BOLT-01 - Hex Bolt")
    assert tab.transaction_table.rowCount() == 2
    assert tab.transaction_table.item(0, 1).text() == "Adjustment Increase"
    assert tab.transaction_table.item(1, 1).text() == "Opening Balance"


def test_requisitions_tab_shows_demand_lines_and_shared_context(qapp, services):
    _enable_inventory_module(services)
    site, storeroom, item, supplier = _create_procurement_context(services)
    register_and_login(services, username_prefix="inventory-ui-req", role_names=("inventory_manager",))
    requisition = services["inventory_procurement_service"].create_requisition(
        requesting_site_id=site.id,
        requesting_storeroom_id=storeroom.id,
        purpose="Restock inspection spares",
    )
    services["inventory_procurement_service"].add_requisition_line(
        requisition.id,
        stock_item_id=item.id,
        quantity_requested=4,
        suggested_supplier_party_id=supplier.id,
        estimated_unit_cost=85.0,
    )

    tab = RequisitionsTab(
        procurement_service=services["inventory_procurement_service"],
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.table.rowCount() == 1
    assert tab.count_badge.text() == "1 requisitions"

    tab.table.selectRow(0)
    qapp.processEvents()

    assert tab.detail_name.text() == requisition.requisition_number
    assert tab.detail_site.text() == f"{site.site_code} - {site.name}"
    assert tab.detail_storeroom.text() == f"{storeroom.storeroom_code} - {storeroom.name}"
    assert tab.lines_table.rowCount() == 1
    assert item.item_code in tab.lines_table.item(0, 1).text()


def test_requisition_and_purchase_order_tabs_enable_draft_edit_actions(qapp, services):
    _enable_inventory_module(services)
    site, storeroom, item, supplier = _create_procurement_context(services)
    register_and_login(services, username_prefix="inventory-ui-edit", role_names=("inventory_manager",))
    procurement = services["inventory_procurement_service"]
    purchasing = services["inventory_purchasing_service"]

    requisition = procurement.create_requisition(
        requesting_site_id=site.id,
        requesting_storeroom_id=storeroom.id,
        purpose="Editable requisition",
    )
    procurement.add_requisition_line(
        requisition.id,
        stock_item_id=item.id,
        quantity_requested=2,
        suggested_supplier_party_id=supplier.id,
    )

    requisitions_tab = RequisitionsTab(
        procurement_service=procurement,
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    requisitions_tab.table.selectRow(0)
    qapp.processEvents()

    assert requisitions_tab.btn_edit.isEnabled()
    assert requisitions_tab.btn_cancel.isEnabled()

    purchase_order = purchasing.create_purchase_order(
        site_id=site.id,
        supplier_party_id=supplier.id,
        currency_code="EUR",
    )
    purchasing.add_purchase_order_line(
        purchase_order.id,
        stock_item_id=item.id,
        destination_storeroom_id=storeroom.id,
        quantity_ordered=2,
        unit_price=45.0,
    )

    purchase_orders_tab = PurchaseOrdersTab(
        purchasing_service=purchasing,
        procurement_service=procurement,
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    purchase_orders_tab.table.selectRow(0)
    qapp.processEvents()

    assert purchase_orders_tab.btn_edit.isEnabled()
    assert purchase_orders_tab.btn_cancel.isEnabled()


def test_inventory_items_tab_can_link_and_unlink_shared_documents_via_ui_actions(qapp, services, monkeypatch):
    _enable_inventory_module(services)
    document = services["document_service"].create_document(
        document_code="DOC-LINK-01",
        title="Inspection Manual",
        document_type="MANUAL",
        storage_kind="REFERENCE",
        storage_uri="vault://docs/inspection-manual",
    )
    register_and_login(services, username_prefix="inventory-ui-docs", role_names=("inventory_manager",))
    item = services["inventory_item_service"].create_item(
        item_code="DOC-LINK-01",
        name="Document Link Item",
        status="ACTIVE",
        stock_uom="EA",
    )

    tab = InventoryItemsTab(
        item_service=services["inventory_item_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    tab.table.selectRow(0)
    qapp.processEvents()

    class _AcceptedDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return QDialog.Accepted

        @property
        def document_id(self):
            return document.id

    monkeypatch.setattr("ui.modules.inventory_procurement.items_tab.InventoryItemDocumentLinkDialog", _AcceptedDialog)

    assert tab.btn_link_document.isEnabled() is True
    assert tab.btn_unlink_document.isEnabled() is True

    tab.link_document()
    qapp.processEvents()

    assert document.title in tab.detail_documents.text()

    tab.unlink_document()
    qapp.processEvents()

    assert tab.detail_documents.text() == "No linked documents"


def test_purchase_orders_tab_shows_line_and_source_requisition(qapp, services):
    _enable_inventory_module(services)
    (
        _site,
        storeroom,
        item,
        supplier,
        requisition,
        purchase_order,
        _purchase_order_line,
        _receipt,
    ) = _create_purchase_flow(services)

    tab = PurchaseOrdersTab(
        purchasing_service=services["inventory_purchasing_service"],
        procurement_service=services["inventory_procurement_service"],
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.table.rowCount() == 1
    assert tab.count_badge.text() == "1 purchase orders"

    tab.table.selectRow(0)
    qapp.processEvents()

    assert tab.detail_name.text() == purchase_order.po_number
    assert tab.detail_supplier.text() == f"{supplier.party_code} - {supplier.party_name}"
    assert tab.detail_source.text() == requisition.requisition_number
    assert tab.lines_table.rowCount() == 1
    assert storeroom.storeroom_code in tab.lines_table.item(0, 2).text()
    assert item.item_code in tab.lines_table.item(0, 1).text()


def test_purchase_orders_tab_enables_send_and_close_for_progressed_orders(qapp, services):
    _enable_inventory_module(services)
    (
        _site,
        _storeroom,
        _item,
        _supplier,
        _requisition,
        purchase_order,
        purchase_order_line,
        _receipt,
    ) = _create_purchase_flow(services)
    purchasing = services["inventory_purchasing_service"]

    tab = PurchaseOrdersTab(
        purchasing_service=purchasing,
        procurement_service=services["inventory_procurement_service"],
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    tab.table.selectRow(0)
    qapp.processEvents()

    assert tab.btn_send.isEnabled() is True
    assert tab.btn_close.isEnabled() is False

    purchasing.send_purchase_order(purchase_order.id)
    purchasing.post_receipt(
        purchase_order.id,
        receipt_lines=[
            {
                "purchase_order_line_id": purchase_order_line.id,
                "quantity_accepted": 5,
                "unit_cost": 115.0,
            }
        ],
    )
    tab.reload_purchase_orders()
    tab.table.selectRow(0)
    qapp.processEvents()

    assert tab.btn_send.isEnabled() is False
    assert tab.btn_close.isEnabled() is True


def test_receiving_tab_shows_outstanding_lines_and_receipt_history(qapp, services):
    _enable_inventory_module(services)
    (
        site,
        _storeroom,
        item,
        _supplier,
        _requisition,
        purchase_order,
        _purchase_order_line,
        receipt,
    ) = _create_purchase_flow(services, with_receipt=True)

    tab = ReceivingTab(
        purchasing_service=services["inventory_purchasing_service"],
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        reference_service=services["inventory_reference_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.work_tabs.count() == 2
    assert tab.work_tabs.tabText(0) == "Outstanding Lines"
    assert tab.work_tabs.tabText(1) == "Receipt History"
    assert tab.table.rowCount() == 1

    tab.table.selectRow(0)
    qapp.processEvents()

    assert tab.detail_name.text() == purchase_order.po_number
    assert tab.detail_site.text() == f"{site.site_code} - {site.name}"
    assert tab.lines_table.rowCount() == 1
    assert tab.receipts_table.rowCount() == 1
    assert item.item_code in tab.lines_table.item(0, 1).text()
    assert tab.receipts_table.item(0, 0).text() == receipt.receipt_number


def test_reservations_tab_shows_active_reservation_and_source_reference(qapp, services):
    _enable_inventory_module(services)
    site, storeroom, item, _supplier = _create_procurement_context(services)
    register_and_login(services, username_prefix="inventory-ui-res", role_names=("inventory_manager",))
    services["inventory_stock_service"].post_opening_balance(
        stock_item_id=item.id,
        storeroom_id=storeroom.id,
        quantity=10.0,
        unit_cost=2.5,
    )
    reservation = services["inventory_reservation_service"].create_reservation(
        stock_item_id=item.id,
        storeroom_id=storeroom.id,
        reserved_qty=4.0,
        source_reference_type="work_order",
        source_reference_id="WO-100",
        notes="Reserve for planned work",
    )

    tab = ReservationsTab(
        reservation_service=services["inventory_reservation_service"],
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.table.rowCount() == 1
    assert tab.count_badge.text() == "1 reservations"

    tab.table.selectRow(0)
    qapp.processEvents()

    assert tab.detail_name.text() == reservation.reservation_number
    assert tab.detail_item.text() == f"{item.item_code} - {item.name}"
    assert tab.detail_storeroom.text() == f"{storeroom.storeroom_code} - {storeroom.name}"
    assert tab.detail_source.text() == "work_order: WO-100"
    assert "Reserved 4.000" in tab.detail_quantities.text()


def test_movements_tab_shows_issue_return_and_transfer_history(qapp, services):
    _enable_inventory_module(services)
    site, source, item, _supplier = _create_procurement_context(services)
    register_and_login(services, username_prefix="inventory-ui-move", role_names=("inventory_manager",))
    destination = services["inventory_service"].create_storeroom(
        storeroom_code="SAT",
        name="Satellite Stores",
        site_id=site.id,
        status="ACTIVE",
    )
    stock = services["inventory_stock_service"]
    stock.post_opening_balance(
        stock_item_id=item.id,
        storeroom_id=source.id,
        quantity=8.0,
        unit_cost=5.5,
    )
    stock.issue_stock(
        stock_item_id=item.id,
        storeroom_id=source.id,
        quantity=2.0,
        reference_type="work_order_issue",
        reference_id="WO-110",
    )
    stock.return_stock(
        stock_item_id=item.id,
        storeroom_id=source.id,
        quantity=1.0,
        reference_type="work_order_return",
        reference_id="WO-110",
    )
    stock.transfer_stock(
        stock_item_id=item.id,
        source_storeroom_id=source.id,
        destination_storeroom_id=destination.id,
        quantity=3.0,
        notes="Move to satellite store",
    )

    tab = MovementsTab(
        stock_service=stock,
        item_service=services["inventory_item_service"],
        inventory_service=services["inventory_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )

    assert tab.table.rowCount() == 4
    assert tab.count_badge.text() == "4 movements"
    assert tab.issue_badge.text() == "1 issues"
    assert tab.transfer_badge.text() == "2 transfers"

    row_types = {tab.table.item(row, 1).text() for row in range(tab.table.rowCount())}
    assert row_types == {"Issue", "Return", "Transfer Out", "Transfer In"}

    for row in range(tab.table.rowCount()):
        if tab.table.item(row, 1).text() == "Transfer Out":
            tab.table.selectRow(row)
            break
    qapp.processEvents()

    assert tab.detail_name.text() == "Transfer Out"
    assert tab.detail_storeroom.text() == f"{source.storeroom_code} - {source.name}"
    assert tab.detail_reference.text().startswith("stock_transfer: INV-TRF-")
    assert "On hand" in tab.detail_result.text()


def test_main_window_exposes_inventory_workspaces_when_module_is_enabled(
    qapp,
    services,
    repo_workspace,
    monkeypatch,
):
    _enable_inventory_module(services)
    register_and_login(services, username_prefix="inventory-shell", role_names=("inventory_manager", "viewer"))
    store = make_settings_store(repo_workspace, prefix="inventory-shell")
    monkeypatch.setattr("ui.platform.shell.main_window.MainWindowSettingsStore", lambda: store)
    monkeypatch.setattr(MainWindow, "_run_startup_update_check", lambda self: None)

    window = MainWindow(services)
    labels = [window.tabs.tabText(i) for i in range(window.tabs.count())]

    assert "Inventory Dashboard" in labels
    assert "Items" in labels
    assert "Storerooms" in labels
    assert "Stock" in labels
    assert "Movements" in labels
    assert "Reservations" in labels
    assert "Requisitions" in labels
    assert "Purchase Orders" in labels
    assert "Receiving" in labels
    assert "Data Exchange" in labels
    assert "Reports" in labels

    inventory_section = window.shell_navigation.tree.topLevelItem(2)
    assert inventory_section.text(0) == "Inventory & Procurement"
    assert _child_labels(inventory_section) == ["Overview", "Master Data", "Operations", "Procurement", "Control"]
    assert _child_labels(inventory_section.child(0)) == ["Inventory Dashboard"]
    assert _child_labels(inventory_section.child(1)) == ["Items", "Storerooms"]
    assert _child_labels(inventory_section.child(2)) == ["Reservations", "Movements", "Stock"]
    assert _child_labels(inventory_section.child(3)) == ["Requisitions", "Purchase Orders", "Receiving"]
    assert _child_labels(inventory_section.child(4)) == ["Data Exchange", "Reports"]


def test_inventory_ui_uses_qdialog_acceptance_constant_in_new_dialog_paths():
    root = Path(__file__).resolve().parents[1]
    inventory_ui = root / "ui" / "modules" / "inventory_procurement"
    texts = [
        (inventory_ui / "item_master" / "items_tab.py").read_text(encoding="utf-8", errors="ignore"),
        (inventory_ui / "procurement" / "purchase_orders_tab.py").read_text(encoding="utf-8", errors="ignore"),
        (inventory_ui / "procurement" / "receiving_tab.py").read_text(encoding="utf-8", errors="ignore"),
        (inventory_ui / "reservation" / "reservations_tab.py").read_text(encoding="utf-8", errors="ignore"),
        (inventory_ui / "procurement" / "requisitions_tab.py").read_text(encoding="utf-8", errors="ignore"),
        (inventory_ui / "inventory" / "storerooms_tab.py").read_text(encoding="utf-8", errors="ignore"),
        (inventory_ui / "stock_control" / "movements_tab.py").read_text(encoding="utf-8", errors="ignore"),
        (inventory_ui / "stock_control" / "stock_tab.py").read_text(encoding="utf-8", errors="ignore"),
    ]

    assert all("dialog.Accepted" not in text for text in texts)
    assert all("QDialog.Accepted" in text for text in texts)


def test_base_stylesheet_explicitly_styles_qplaintextedit_note_surfaces():
    css = base_stylesheet()

    assert "QPlainTextEdit" in css
    assert "QPlainTextEdit:focus" in css
