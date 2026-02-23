from __future__ import annotations

from datetime import date
from pathlib import Path

from matplotlib.axes import Axes
from matplotlib.dates import date2num
from openpyxl import load_workbook

from core.models import CostType, DependencyType
from core.exceptions import BusinessRuleError
from core.reporting import api as reporting_api
from core.services.reporting.models import (
    CostBreakdownRow,
    EarnedValueMetrics,
    EvmSeriesPoint,
    GanttTaskBar,
    ProjectKPI,
    ResourceLoadRow,
    TaskVarianceRow,
)


def _setup_report_project(services):
    ps = services["project_service"]
    ts = services["task_service"]
    rs = services["resource_service"]
    cs = services["cost_service"]
    bs = services["baseline_service"]

    project = ps.create_project(
        "Export Validation Project",
        "",
        start_date=date(2023, 11, 6),
        end_date=date(2023, 11, 30),
        currency="USD",
    )
    pid = project.id

    t1 = ts.create_task(pid, "Task Alpha", start_date=date(2023, 11, 6), duration_days=2)
    t2 = ts.create_task(pid, "Task Beta", duration_days=2)
    ts.add_dependency(t1.id, t2.id, DependencyType.FINISH_TO_START, lag_days=0)

    res = rs.create_resource("Exporter Dev", "Developer", hourly_rate=100.0, currency_code="USD")
    assignment = ts.assign_resource(t1.id, res.id, allocation_percent=50.0)
    ts.set_assignment_hours(assignment.id, 4.0)

    cs.add_cost_item(
        project_id=pid,
        task_id=t1.id,
        description="Material line",
        planned_amount=120.0,
        actual_amount=60.0,
        cost_type=CostType.MATERIAL,
        currency_code="USD",
    )

    baseline = bs.create_baseline(pid, "Baseline Export")
    ts.update_progress(t1.id, percent_complete=50.0)
    return pid, baseline.id


def test_gantt_export_inclusive_duration_for_one_day_tasks(services, tmp_path, monkeypatch):
    ps = services["project_service"]
    ts = services["task_service"]

    project = ps.create_project("Gantt One Day", "")
    pid = project.id
    ts.create_task(pid, "One Day Task", start_date=date(2023, 11, 6), duration_days=1)

    widths = []
    original_barh = Axes.barh

    def _spy_barh(self, *args, **kwargs):
        width = args[1] if len(args) > 1 else kwargs.get("width")
        widths.append(width)
        return original_barh(self, *args, **kwargs)

    monkeypatch.setattr(Axes, "barh", _spy_barh)

    output = tmp_path / "gantt.png"
    reporting_api.generate_gantt_png(services["reporting_service"], pid, output)

    assert output.exists()
    assert output.stat().st_size > 0
    assert any(float(w) >= 0.5 for w in widths if w is not None)


def test_gantt_render_uses_day_cell_alignment_for_inclusive_finish(services, tmp_path, monkeypatch):
    ps = services["project_service"]
    ts = services["task_service"]
    rp = services["reporting_service"]

    project = ps.create_project("Gantt Day Cell Alignment", "")
    pid = project.id
    task = ts.create_task(pid, "Aligned Task", start_date=date(2023, 11, 6), duration_days=3)

    calls = []
    original_barh = Axes.barh

    def _spy_barh(self, *args, **kwargs):
        calls.append((args, kwargs))
        return original_barh(self, *args, **kwargs)

    monkeypatch.setattr(Axes, "barh", _spy_barh)

    output = tmp_path / "gantt_aligned.png"
    reporting_api.generate_gantt_png(services["reporting_service"], pid, output)

    assert output.exists()
    bars = rp.get_gantt_data(pid)
    bar = next(b for b in bars if b.task_id == task.id)

    # Main bars have dark edge color; progress overlays do not.
    main = next(
        (entry for entry in calls if entry[1].get("edgecolor") == "#0F172A"),
        None,
    )
    assert main is not None
    args, kwargs = main
    width = float(args[1])
    left = float(kwargs.get("left"))

    right = left + width
    start_num = date2num(bar.start)
    end_num = date2num(bar.end)

    # Point-to-point rendering: bar starts at start tick and ends near end tick.
    assert start_num <= left < (start_num + 0.2)
    assert (end_num - 0.05) <= right <= (end_num + 0.2)


def test_evm_export_png_is_skipped_when_no_series_data(tmp_path):
    class EmptySeriesService:
        def get_evm_series(self, _project_id, baseline_id=None, as_of=None):
            return []

    output = tmp_path / "evm_empty.png"
    result = reporting_api.generate_evm_png(EmptySeriesService(), "p1", output)

    assert result == output
    assert not output.exists()


def test_evm_export_png_is_skipped_when_baseline_is_missing(tmp_path):
    class NoBaselineService:
        def get_evm_series(self, _project_id, baseline_id=None, as_of=None):
            raise BusinessRuleError("No baseline found.", code="NO_BASELINE")

    output = tmp_path / "evm_no_baseline.png"
    result = reporting_api.generate_evm_png(NoBaselineService(), "p1", output)

    assert result == output
    assert not output.exists()


def test_evm_export_png_generates_image_when_series_exists(tmp_path):
    class SeriesService:
        def get_evm_series(self, _project_id, baseline_id=None, as_of=None):
            return [
                EvmSeriesPoint(
                    period_end=date(2023, 11, 30),
                    PV=100.0,
                    EV=80.0,
                    AC=90.0,
                    BAC=120.0,
                    CPI=0.89,
                    SPI=0.80,
                ),
                EvmSeriesPoint(
                    period_end=date(2023, 12, 31),
                    PV=120.0,
                    EV=110.0,
                    AC=115.0,
                    BAC=120.0,
                    CPI=0.96,
                    SPI=0.92,
                ),
            ]

    output = tmp_path / "evm.png"
    result = reporting_api.generate_evm_png(SeriesService(), "p1", output)

    assert result == output
    assert output.exists()
    assert output.stat().st_size > 0


def test_excel_export_contains_expected_sections_when_baseline_exists(services, tmp_path):
    pid, baseline_id = _setup_report_project(services)
    output = tmp_path / "report.xlsx"

    reporting_api.generate_excel_report(
        services["reporting_service"],
        pid,
        output,
        baseline_id=baseline_id,
        as_of=date(2023, 11, 30),
    )

    wb = load_workbook(output)
    names = set(wb.sheetnames)
    assert {"Overview", "Tasks", "Resources", "EVM", "Variance", "Cost Breakdown"}.issubset(names)
    assert wb["Overview"]["A1"].value.startswith("Project KPIs - ")
    assert wb["Tasks"]["A1"].value == "Task ID"
    assert wb["EVM"]["A2"].value == "Metric"
    assert wb["EVM"]["D2"].value == "Period End"
    assert wb["Cost Breakdown"]["A1"].value == "Type"


def test_excel_export_without_baseline_skips_evm_sheet(services, tmp_path):
    ps = services["project_service"]
    ts = services["task_service"]

    project = ps.create_project("No Baseline Export", "")
    pid = project.id
    ts.create_task(pid, "Task No Base", start_date=date(2023, 11, 6), duration_days=2)

    output = tmp_path / "report_no_baseline.xlsx"
    reporting_api.generate_excel_report(services["reporting_service"], pid, output)

    wb = load_workbook(output)
    assert "EVM" not in wb.sheetnames
    assert {"Overview", "Tasks", "Resources"}.issubset(set(wb.sheetnames))


def test_pdf_export_succeeds_when_gantt_generation_fails(services, tmp_path, monkeypatch):
    pid, _baseline_id = _setup_report_project(services)

    def _raise_gantt(*_args, **_kwargs):
        raise ValueError("No tasks with dates available for Gantt chart.")

    monkeypatch.setattr("core.reporting.api.generate_gantt_png", _raise_gantt)

    output = tmp_path / "report.pdf"
    reporting_api.generate_pdf_report(
        services["reporting_service"],
        pid,
        output,
        temp_dir=tmp_path / "tmp_reports",
        as_of=date(2023, 11, 30),
    )

    assert output.exists()
    assert output.read_bytes().startswith(b"%PDF")


def test_reporting_api_populates_optional_contexts(monkeypatch, tmp_path):
    kpi = ProjectKPI(
        project_id="p1",
        name="API Context",
        start_date=date(2023, 11, 6),
        end_date=date(2023, 11, 30),
        duration_working_days=19,
        tasks_total=1,
        tasks_completed=0,
        tasks_in_progress=1,
        task_blocked=0,
        tasks_not_started=0,
        critical_tasks=1,
        late_tasks=0,
        total_planned_cost=100.0,
        total_actual_cost=50.0,
        cost_variance=-50.0,
        total_committed_cost=60.0,
        committment_variance=-40.0,
    )
    bars = [
        GanttTaskBar(
            task_id="t1",
            name="Task API",
            start=date(2023, 11, 6),
            end=date(2023, 11, 7),
            is_critical=True,
            percent_complete=50.0,
            status="IN_PROGRESS",
        )
    ]
    resources = [ResourceLoadRow(resource_id="r1", resource_name="Res", total_allocation_percent=50.0, tasks_count=1)]
    evm = EarnedValueMetrics(
        as_of=date(2023, 11, 30),
        baseline_id="b1",
        BAC=100.0,
        PV=80.0,
        EV=60.0,
        AC=50.0,
        CPI=1.2,
        SPI=0.75,
        EAC=83.33,
        ETC=33.33,
        VAC=16.67,
    )
    series = [
        EvmSeriesPoint(
            period_end=date(2023, 11, 30),
            PV=80.0,
            EV=60.0,
            AC=50.0,
            BAC=100.0,
            CPI=1.2,
            SPI=0.75,
        )
    ]
    variance = [
        TaskVarianceRow(
            task_id="t1",
            task_name="Task API",
            baseline_start=date(2023, 11, 6),
            baseline_finish=date(2023, 11, 7),
            current_start=date(2023, 11, 6),
            current_finish=date(2023, 11, 8),
            start_variance_days=0,
            finish_variance_days=1,
            is_critical=True,
        )
    ]
    cost_breakdown = [
        CostBreakdownRow(cost_type="MATERIAL", currency="USD", planned=100.0, actual=50.0)
    ]

    class DummyReportingService:
        def get_project_kpis(self, _project_id):
            return kpi

        def get_gantt_data(self, _project_id):
            return bars

        def get_resource_load_summary(self, _project_id):
            return resources

        def get_earned_value(self, _project_id, baseline_id=None, as_of=None):
            assert baseline_id == "b1"
            assert as_of == date(2023, 11, 30)
            return evm

        def get_evm_series(self, _project_id, baseline_id=None, as_of=None):
            assert baseline_id == "b1"
            assert as_of == date(2023, 11, 30)
            return series

        def get_baseline_schedule_variance(self, _project_id, baseline_id=None):
            assert baseline_id == "b1"
            return variance

        def get_cost_breakdown(self, _project_id, as_of=None, baseline_id=None):
            assert baseline_id == "b1"
            assert as_of == date(2023, 11, 30)
            return cost_breakdown

    captured = {}

    class _FakeExcelRenderer:
        def render(self, ctx, output_path):
            captured["excel_ctx"] = ctx
            captured["excel_path"] = output_path
            return output_path

    class _FakePdfRenderer:
        def render(self, ctx, output_path):
            captured["pdf_ctx"] = ctx
            captured["pdf_path"] = output_path
            return output_path

    monkeypatch.setattr(reporting_api, "ExcelReportRenderer", lambda: _FakeExcelRenderer())
    monkeypatch.setattr(reporting_api, "PdfReportRenderer", lambda: _FakePdfRenderer())
    monkeypatch.setattr(reporting_api, "generate_gantt_png", lambda *_args, **_kwargs: Path(tmp_path / "gantt.png"))

    service = DummyReportingService()
    as_of = date(2023, 11, 30)

    reporting_api.generate_excel_report(service, "p1", tmp_path / "api.xlsx", baseline_id="b1", as_of=as_of)
    reporting_api.generate_pdf_report(service, "p1", tmp_path / "api.pdf", baseline_id="b1", as_of=as_of)

    assert captured["excel_ctx"].evm is evm
    assert captured["excel_ctx"].evm_series == series
    assert captured["excel_ctx"].baseline_variance == variance
    assert captured["excel_ctx"].cost_breakdown == cost_breakdown
    assert captured["excel_ctx"].as_of == as_of

    assert captured["pdf_ctx"].evm is evm
    assert captured["pdf_ctx"].evm_series == series
    assert captured["pdf_ctx"].baseline_variance == variance
    assert captured["pdf_ctx"].cost_breakdown == cost_breakdown
    assert captured["pdf_ctx"].as_of == as_of
