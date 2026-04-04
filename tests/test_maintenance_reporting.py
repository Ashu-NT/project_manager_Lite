from __future__ import annotations

from datetime import datetime, timedelta, timezone

import pytest
from openpyxl import load_workbook

from core.platform.auth.session import UserSessionPrincipal
from core.platform.common.exceptions import BusinessRuleError


class _EnabledMaintenanceModuleCatalog:
    def get_entitlement(self, _module_code: str):
        return type(
            "Entitlement",
            (),
            {
                "label": "Maintenance Management",
                "runtime_enabled": True,
                "lifecycle_status": "active",
            },
        )()


def _create_reporting_context(services):
    services["maintenance_reporting_service"]._module_catalog_service = _EnabledMaintenanceModuleCatalog()
    site = services["site_service"].create_site(
        site_code="MNT-REP",
        name="Maintenance Reporting Site",
        currency_code="EUR",
    )
    location = services["maintenance_location_service"].create_location(
        site_id=site.id,
        location_code="REP-AREA",
        name="Reporting Area",
    )
    asset = services["maintenance_asset_service"].create_asset(
        site_id=site.id,
        location_id=location.id,
        asset_code="REP-ASSET",
        name="Reporting Asset",
    )
    integration_source = services["maintenance_integration_source_service"].create_source(
        integration_code="REP-IOT",
        name="Reporting Gateway",
        integration_type="iot_gateway",
        endpoint_or_path="mqtt://reporting/gateway",
    )
    sensor = services["maintenance_sensor_service"].create_sensor(
        site_id=site.id,
        sensor_code="REP-SENSOR",
        sensor_name="Reporting Run Hours",
        asset_id=asset.id,
        sensor_type="running_hours",
        source_type="iot_gateway",
        source_name=integration_source.name,
        source_key="reporting.asset.hours",
        unit="H",
    )
    services["maintenance_sensor_exception_service"].raise_exception(
        exception_type="stale_reading",
        sensor_id=sensor.id,
        integration_source_id=integration_source.id,
        source_batch_id="REP-BATCH-001",
        message="Reporting sensor feed is stale.",
        notes="Raised for the reliability exception review workbook.",
    )
    symptom = services["maintenance_failure_code_service"].create_failure_code(
        failure_code="seal-leak",
        name="Seal Leak",
        code_type="symptom",
    )
    cause = services["maintenance_failure_code_service"].create_failure_code(
        failure_code="misalignment",
        name="Misalignment",
        code_type="cause",
    )

    now = datetime.now(timezone.utc)

    backlog_request = services["maintenance_work_request_service"].create_work_request(
        site_id=site.id,
        work_request_code="WR-REP-OPEN",
        source_type="manual",
        request_type="breakdown",
        asset_id=asset.id,
        location_id=location.id,
        title="Open corrective work",
        failure_symptom_code=symptom.failure_code,
    )
    backlog_order = services["maintenance_work_order_service"].create_work_order(
        site_id=site.id,
        work_order_code="WO-REP-OPEN",
        work_order_type="corrective",
        source_type="work_request",
        source_id=backlog_request.id,
    )
    backlog_order = services["maintenance_work_order_service"].update_work_order(
        backlog_order.id,
        status="planned",
        planned_start=now - timedelta(days=2),
        planned_end=now - timedelta(days=1),
        expected_version=backlog_order.version,
    )

    preventive_request = services["maintenance_work_request_service"].create_work_request(
        site_id=site.id,
        work_request_code="WR-REP-PM",
        source_type="manual",
        request_type="preventive",
        asset_id=asset.id,
        location_id=location.id,
        title="Completed preventive work",
        failure_symptom_code=symptom.failure_code,
    )
    preventive_order = services["maintenance_work_order_service"].create_work_order(
        site_id=site.id,
        work_order_code="WO-REP-PM",
        work_order_type="preventive",
        source_type="work_request",
        source_id=preventive_request.id,
        is_preventive=True,
    )
    preventive_order = services["maintenance_work_order_service"].update_work_order(
        preventive_order.id,
        status="planned",
        planned_start=now - timedelta(days=6),
        planned_end=now - timedelta(days=5),
        expected_version=preventive_order.version,
    )
    preventive_order = services["maintenance_work_order_service"].update_work_order(
        preventive_order.id,
        status="released",
        expected_version=preventive_order.version,
    )
    preventive_order = services["maintenance_work_order_service"].update_work_order(
        preventive_order.id,
        status="in_progress",
        expected_version=preventive_order.version,
    )
    preventive_order = services["maintenance_work_order_service"].update_work_order(
        preventive_order.id,
        status="completed",
        failure_code=symptom.failure_code,
        root_cause_code=cause.failure_code,
        expected_version=preventive_order.version,
    )
    services["maintenance_downtime_event_service"].create_downtime_event(
        work_order_id=preventive_order.id,
        started_at=(now - timedelta(days=5, hours=2)).isoformat(),
        ended_at=(now - timedelta(days=5, hours=1)).isoformat(),
        downtime_type="unplanned",
        reason_code=symptom.failure_code,
        impact_notes="Downtime during corrective intervention.",
    )

    repeat_request = services["maintenance_work_request_service"].create_work_request(
        site_id=site.id,
        work_request_code="WR-REP-REPEAT",
        source_type="manual",
        request_type="breakdown",
        asset_id=asset.id,
        location_id=location.id,
        title="Repeat corrective work",
        failure_symptom_code=symptom.failure_code,
    )
    repeat_order = services["maintenance_work_order_service"].create_work_order(
        site_id=site.id,
        work_order_code="WO-REP-REPEAT",
        work_order_type="corrective",
        source_type="work_request",
        source_id=repeat_request.id,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="planned",
        planned_start=now - timedelta(days=12),
        planned_end=now - timedelta(days=11),
        expected_version=repeat_order.version,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="released",
        expected_version=repeat_order.version,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="in_progress",
        expected_version=repeat_order.version,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="completed",
        failure_code=symptom.failure_code,
        root_cause_code=cause.failure_code,
        downtime_minutes=75,
        expected_version=repeat_order.version,
    )
    services["maintenance_downtime_event_service"].create_downtime_event(
        work_order_id=repeat_order.id,
        started_at=(now - timedelta(days=11, hours=2)).isoformat(),
        ended_at=(now - timedelta(days=11, hours=1)).isoformat(),
        downtime_type="unplanned",
        reason_code=symptom.failure_code,
        impact_notes="Repeat downtime event for recurring analysis.",
    )

    return site, asset, symptom, cause


def test_maintenance_reporting_service_generates_report_packs(services, tmp_path):
    site, asset, symptom, _cause = _create_reporting_context(services)
    reporting = services["maintenance_reporting_service"]

    backlog_excel = reporting.generate_backlog_excel(tmp_path / "maintenance-backlog.xlsx", site_id=site.id, asset_id=asset.id, days=365)
    downtime_pdf = reporting.generate_downtime_pdf(tmp_path / "maintenance-downtime.pdf", site_id=site.id, asset_id=asset.id, days=365)
    pm_excel = reporting.generate_pm_compliance_excel(tmp_path / "maintenance-pm.xlsx", site_id=site.id, asset_id=asset.id, days=365)
    execution_pdf = reporting.generate_execution_overview_pdf(tmp_path / "maintenance-execution.pdf", site_id=site.id, asset_id=asset.id, days=365)
    recurring_excel = reporting.generate_recurring_failures_excel(tmp_path / "maintenance-recurring.xlsx", site_id=site.id, asset_id=asset.id, days=365)
    exception_excel = reporting.generate_exception_review_excel(tmp_path / "maintenance-exceptions.xlsx", site_id=site.id, asset_id=asset.id, days=365)

    workbook = load_workbook(backlog_excel.file_path)
    pm_workbook = load_workbook(pm_excel.file_path)
    recurring_workbook = load_workbook(recurring_excel.file_path)
    exception_workbook = load_workbook(exception_excel.file_path)

    assert backlog_excel.file_path.exists()
    assert downtime_pdf.file_path.exists()
    assert execution_pdf.file_path.exists()
    assert recurring_excel.file_path.exists()
    assert exception_excel.file_path.exists()
    assert downtime_pdf.file_path.stat().st_size > 0
    assert execution_pdf.file_path.stat().st_size > 0
    assert {"Summary", "Open Work Orders", "Top Root Causes", "Recurring Failure Patterns"} <= set(workbook.sheetnames)
    assert {"Summary", "Preventive Work Orders"} <= set(pm_workbook.sheetnames)
    assert {"Summary", "Leading Root Causes", "Recurring Failure Review"} <= set(recurring_workbook.sheetnames)
    assert {"Summary", "Open Sensor Exceptions", "Recurring Failure Signals"} <= set(exception_workbook.sheetnames)
    assert workbook["Open Work Orders"]["A4"].value == "WO-REP-OPEN"
    assert recurring_workbook["Recurring Failure Review"]["A4"].value == "Reporting Asset"
    assert exception_workbook["Open Sensor Exceptions"]["D4"].value == "REP-SENSOR - Reporting Run Hours"
    root_cause_sheet_values = {str(workbook["Top Root Causes"][row][0].value) for row in range(4, 8) if workbook["Top Root Causes"][row][0].value}
    assert "Seal Leak" in root_cause_sheet_values
    assert symptom.failure_code == "SEAL-LEAK"


def test_maintenance_reporting_service_requires_runtime_permissions(services, tmp_path):
    _create_reporting_context(services)
    services["user_session"].set_principal(
        UserSessionPrincipal(
            user_id="u-maint-read",
            username="maint-read-only",
            display_name="Maintenance Read Only",
            role_names=frozenset({"viewer"}),
            permissions=frozenset({"maintenance.read", "report.view"}),
        )
    )

    with pytest.raises(BusinessRuleError, match="Permission denied") as exc:
        services["maintenance_reporting_service"].generate_backlog_excel(tmp_path / "denied-maintenance.xlsx")

    assert exc.value.code == "PERMISSION_DENIED"
