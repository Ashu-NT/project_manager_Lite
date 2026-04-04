from __future__ import annotations

from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtWidgets import QDialog

from tests.ui_runtime_helpers import make_settings_store
from ui.modules.maintenance_management import (
    MaintenanceAssetsTab,
    MaintenanceDashboardTab,
    MaintenanceDocumentsTab,
    MaintenanceLocationsTab,
    MaintenancePlannerTab,
    MaintenancePreventivePlansTab,
    MaintenanceReliabilityTab,
    MaintenanceSensorsTab,
    MaintenanceSystemsTab,
    MaintenanceWorkOrdersTab,
)
from ui.platform.shell.main_window import MainWindow


def _child_labels(item) -> list[str]:
    return [item.child(i).text(0) for i in range(item.childCount())]


def _top_level_labels(tree) -> list[str]:
    return [tree.topLevelItem(i).text(0) for i in range(tree.topLevelItemCount())]


def _enable_maintenance_module(services) -> None:
    services["module_catalog_service"].set_module_state(
        "maintenance_management",
        licensed=True,
        enabled=True,
        lifecycle_status="active",
    )


def _select_combo_value(combo, value) -> None:
    index = combo.findData(value)
    assert index >= 0
    combo.setCurrentIndex(index)


def _find_row_by_contains(table, column: int, needle: str) -> int:
    for row in range(table.rowCount()):
        item = table.item(row, column)
        if item is not None and needle in item.text():
            return row
    return -1


def _mute_message_boxes(monkeypatch) -> None:
    monkeypatch.setattr("PySide6.QtWidgets.QMessageBox.information", lambda *args, **kwargs: None)
    monkeypatch.setattr("PySide6.QtWidgets.QMessageBox.warning", lambda *args, **kwargs: None)
    monkeypatch.setattr("PySide6.QtWidgets.QMessageBox.critical", lambda *args, **kwargs: None)


def _create_maintenance_context(services):
    site = services["site_service"].create_site(
        site_code="MNT-UI",
        name="Maintenance UI Site",
        currency_code="EUR",
    )
    location = services["maintenance_location_service"].create_location(
        site_id=site.id,
        location_code="area-a",
        name="Area A",
    )
    system = services["maintenance_system_service"].create_system(
        site_id=site.id,
        location_id=location.id,
        system_code="pump-line",
        name="Pump Line",
    )
    asset = services["maintenance_asset_service"].create_asset(
        site_id=site.id,
        location_id=location.id,
        system_id=system.id,
        asset_code="pump-101",
        name="Pump 101",
    )
    services["maintenance_asset_component_service"].create_component(
        asset_id=asset.id,
        component_code="seal-001",
        name="Mechanical Seal",
        component_type="seal",
        is_critical_component=True,
    )
    source = services["maintenance_integration_source_service"].create_source(
        integration_code="iot-ui-1",
        name="IoT Gateway UI",
        integration_type="iot_gateway",
        endpoint_or_path="mqtt://broker/pumps",
        authentication_mode="token",
        schedule_expression="*/10 * * * *",
    )
    sensor = services["maintenance_sensor_service"].create_sensor(
        site_id=site.id,
        sensor_code="run-hours-101",
        sensor_name="Pump 101 Running Hours",
        asset_id=asset.id,
        sensor_type="running_hours",
        source_type="iot_gateway",
        source_name=source.name,
        source_key="pump-101.hours",
        unit="H",
    )
    services["maintenance_sensor_source_mapping_service"].create_mapping(
        integration_source_id=source.id,
        sensor_id=sensor.id,
        external_equipment_key="PUMP-101",
        external_measurement_key="run_hours",
    )
    services["maintenance_sensor_reading_service"].record_reading(
        sensor_id=sensor.id,
        reading_value="1240.5",
        reading_unit="H",
        quality_state="stale",
        source_name=source.name,
        source_batch_id="BATCH-UI-001",
    )
    services["maintenance_integration_source_service"].record_sync_failure(
        source.id,
        error_message="Gateway timeout during maintenance UI refresh.",
        expected_version=source.version,
    )
    task_template = services["maintenance_task_template_service"].create_task_template(
        task_template_code="pm-seal-check",
        name="Seal inspection route",
        maintenance_type="preventive",
        template_status="active",
        estimated_minutes=60,
        required_skill="Mechanical",
    )
    services["maintenance_task_step_template_service"].create_step_template(
        task_template_id=task_template.id,
        step_number=1,
        instruction="Inspect the seal assembly and capture abnormal wear indicators.",
        expected_result="Seal condition confirmed and documented.",
        requires_confirmation=True,
    )
    symptom = services["maintenance_failure_code_service"].create_failure_code(
        failure_code="seal-leak",
        name="Seal Leak",
        code_type="symptom",
    )
    cause = services["maintenance_failure_code_service"].create_failure_code(
        failure_code="misalignment",
        name="Misalignment",
        code_type="cause",
    )
    now = datetime.now(timezone.utc)
    assigned_employee = services["employee_service"].create_employee(
        employee_code="mnt-tech-01",
        full_name="Maintenance Technician 01",
        site_id=site.id,
        site_name=site.name,
        department="Maintenance",
        title="Technician",
    )
    request_open = services["maintenance_work_request_service"].create_work_request(
        site_id=site.id,
        work_request_code="wr-ui-open",
        source_type="manual",
        request_type="breakdown",
        asset_id=asset.id,
        location_id=location.id,
        system_id=system.id,
        title="Open seal leak",
        failure_symptom_code=symptom.failure_code,
    )
    request_open = services["maintenance_work_request_service"].update_work_request(
        request_open.id,
        status="TRIAGED",
        expected_version=request_open.version,
    )
    open_order = services["maintenance_work_order_service"].create_work_order(
        site_id=site.id,
        work_order_code="wo-ui-open",
        work_order_type="corrective",
        source_type="work_request",
        source_id=request_open.id,
    )
    open_order = services["maintenance_work_order_service"].update_work_order(
        open_order.id,
        status="planned",
        assigned_employee_id=assigned_employee.id,
        planned_start=now - timedelta(days=2),
        planned_end=now - timedelta(days=1),
        expected_version=open_order.version,
    )
    open_task = services["maintenance_work_order_task_service"].create_task(
        work_order_id=open_order.id,
        task_name="Inspect seal housing",
        assigned_employee_id=assigned_employee.id,
        required_skill="Mechanical",
        estimated_minutes=45,
        completion_rule="all_steps_required",
    )
    services["maintenance_work_order_task_step_service"].create_step(
        work_order_task_id=open_task.id,
        instruction="Verify isolation, inspect seal area, and capture readings.",
        expected_result="Seal area inspected and condition confirmed.",
        requires_confirmation=True,
        requires_measurement=True,
        measurement_unit="C",
    )
    services["maintenance_work_order_material_requirement_service"].create_requirement(
        work_order_id=open_order.id,
        description="Seal kit",
        required_qty="1",
        required_uom="EA",
        is_stock_item=False,
        notes="Direct-charge seal kit pending release.",
    )

    request_closed = services["maintenance_work_request_service"].create_work_request(
        site_id=site.id,
        work_request_code="wr-ui-closed",
        source_type="manual",
        request_type="breakdown",
        asset_id=asset.id,
        location_id=location.id,
        system_id=system.id,
        title="Closed leak history",
        failure_symptom_code=symptom.failure_code,
    )
    closed_order = services["maintenance_work_order_service"].create_work_order(
        site_id=site.id,
        work_order_code="wo-ui-closed",
        work_order_type="corrective",
        source_type="work_request",
        source_id=request_closed.id,
    )
    request_closed = services["maintenance_work_request_service"].update_work_request(
        request_closed.id,
        status="CONVERTED",
        expected_version=request_closed.version,
    )
    closed_order = services["maintenance_work_order_service"].update_work_order(
        closed_order.id,
        status="planned",
        planned_start=now - timedelta(days=10),
        planned_end=now - timedelta(days=9),
        expected_version=closed_order.version,
    )
    closed_order = services["maintenance_work_order_service"].update_work_order(
        closed_order.id,
        status="released",
        expected_version=closed_order.version,
    )
    closed_order = services["maintenance_work_order_service"].update_work_order(
        closed_order.id,
        status="in_progress",
        expected_version=closed_order.version,
    )
    closed_order = services["maintenance_work_order_service"].update_work_order(
        closed_order.id,
        status="completed",
        failure_code=symptom.failure_code,
        root_cause_code=cause.failure_code,
        downtime_minutes=90,
        expected_version=closed_order.version,
    )
    services["maintenance_downtime_event_service"].create_downtime_event(
        work_order_id=closed_order.id,
        started_at=(now - timedelta(days=10, hours=2)).isoformat(),
        ended_at=(now - timedelta(days=10, hours=1)).isoformat(),
        downtime_type="unplanned",
        reason_code=symptom.failure_code,
        impact_notes="Unplanned production loss.",
    )
    request_repeat = services["maintenance_work_request_service"].create_work_request(
        site_id=site.id,
        work_request_code="wr-ui-repeat",
        source_type="manual",
        request_type="breakdown",
        asset_id=asset.id,
        location_id=location.id,
        system_id=system.id,
        title="Repeat leak history",
        failure_symptom_code=symptom.failure_code,
    )
    repeat_order = services["maintenance_work_order_service"].create_work_order(
        site_id=site.id,
        work_order_code="wo-ui-repeat",
        work_order_type="corrective",
        source_type="work_request",
        source_id=request_repeat.id,
    )
    request_repeat = services["maintenance_work_request_service"].update_work_request(
        request_repeat.id,
        status="CONVERTED",
        expected_version=request_repeat.version,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="planned",
        planned_start=now - timedelta(days=20),
        planned_end=now - timedelta(days=19),
        expected_version=repeat_order.version,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="released",
        expected_version=repeat_order.version,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="in_progress",
        expected_version=repeat_order.version,
    )
    repeat_order = services["maintenance_work_order_service"].update_work_order(
        repeat_order.id,
        status="completed",
        failure_code=symptom.failure_code,
        root_cause_code=cause.failure_code,
        downtime_minutes=45,
        expected_version=repeat_order.version,
    )
    services["maintenance_downtime_event_service"].create_downtime_event(
        work_order_id=repeat_order.id,
        started_at=(now - timedelta(days=20, hours=1)).isoformat(),
        ended_at=(now - timedelta(days=20, minutes=15)).isoformat(),
        downtime_type="unplanned",
        reason_code=symptom.failure_code,
        impact_notes="Repeat leak recurrence.",
    )
    request_deferred = services["maintenance_work_request_service"].create_work_request(
        site_id=site.id,
        work_request_code="wr-ui-deferred",
        source_type="manual",
        request_type="inspection",
        asset_id=asset.id,
        location_id=location.id,
        system_id=system.id,
        title="Deferred inspection follow-up",
        failure_symptom_code=symptom.failure_code,
    )
    services["maintenance_work_request_service"].update_work_request(
        request_deferred.id,
        status="DEFERRED",
        expected_version=request_deferred.version,
    )
    due_plan = services["maintenance_preventive_plan_service"].create_preventive_plan(
        site_id=site.id,
        plan_code="pm-ui-due",
        name="Due seal inspection",
        asset_id=asset.id,
        plan_type="preventive",
        priority="high",
        trigger_mode="calendar",
        calendar_frequency_unit="weekly",
        calendar_frequency_value=1,
        next_due_at=(now - timedelta(days=1)).isoformat(),
        auto_generate_work_order=True,
        status="active",
    )
    services["maintenance_preventive_plan_task_service"].create_plan_task(
        plan_id=due_plan.id,
        task_template_id=task_template.id,
        sequence_no=1,
        trigger_scope="inherit_plan",
        estimated_minutes_override=50,
    )
    due_soon_plan = services["maintenance_preventive_plan_service"].create_preventive_plan(
        site_id=site.id,
        plan_code="pm-ui-soon",
        name="Due soon seal review",
        asset_id=asset.id,
        plan_type="inspection",
        priority="medium",
        trigger_mode="calendar",
        calendar_frequency_unit="monthly",
        calendar_frequency_value=1,
        next_due_at=(now + timedelta(days=7)).isoformat(),
        status="active",
    )
    services["maintenance_preventive_plan_task_service"].create_plan_task(
        plan_id=due_soon_plan.id,
        task_template_id=task_template.id,
        sequence_no=1,
        trigger_scope="inherit_plan",
    )
    blocked_plan = services["maintenance_preventive_plan_service"].create_preventive_plan(
        site_id=site.id,
        plan_code="pm-ui-blocked",
        name="Blocked hybrid inspection",
        asset_id=asset.id,
        plan_type="condition_based",
        priority="high",
        trigger_mode="hybrid",
        calendar_frequency_unit="monthly",
        calendar_frequency_value=1,
        sensor_id=sensor.id,
        sensor_threshold="1200",
        sensor_direction="greater_or_equal",
        next_due_at=(now + timedelta(days=14)).isoformat(),
        status="active",
    )
    services["maintenance_preventive_plan_task_service"].create_plan_task(
        plan_id=blocked_plan.id,
        task_template_id=task_template.id,
        sequence_no=1,
        trigger_scope="task_override",
        trigger_mode_override="sensor",
        sensor_id_override=sensor.id,
        sensor_threshold_override="1200",
        sensor_direction_override="greater_or_equal",
    )
    services["maintenance_preventive_plan_service"].create_preventive_plan(
        site_id=site.id,
        plan_code="pm-ui-paused",
        name="Paused overhaul library",
        system_id=system.id,
        plan_type="preventive",
        priority="low",
        trigger_mode="calendar",
        calendar_frequency_unit="quarterly",
        calendar_frequency_value=1,
        next_due_at=(now + timedelta(days=45)).isoformat(),
        status="paused",
        is_active=False,
    )
    return site, location, system, asset, symptom


def test_maintenance_locations_tab_supports_create_edit_and_toggle(qapp, services, monkeypatch):
    _enable_maintenance_module(services)
    _mute_message_boxes(monkeypatch)
    site, location, _system, _asset, _symptom = _create_maintenance_context(services)

    tab = MaintenanceLocationsTab(
        location_service=services["maintenance_location_service"],
        site_service=services["site_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.table.rowCount() >= 1
    assert "AREA-A" in tab.table.item(0, 0).text()

    class FakeCreateDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return QDialog.DialogCode.Accepted

        @property
        def site_id(self):
            return site.id

        @property
        def location_code(self):
            return "area-b"

        @property
        def name(self):
            return "Area B"

        @property
        def parent_location_id(self):
            return location.id

        @property
        def location_type(self):
            return "AREA"

        @property
        def criticality(self):
            return "HIGH"

        @property
        def status(self):
            return "ACTIVE"

        @property
        def description(self):
            return "Secondary maintenance area"

        @property
        def notes(self):
            return "Created from maintenance libraries UI."

        @property
        def is_active(self):
            return True

    monkeypatch.setattr(
        "ui.modules.maintenance_management.locations.tab.MaintenanceLocationEditDialog",
        FakeCreateDialog,
    )
    tab.btn_new_location.click()
    qapp.processEvents()
    created_row = _find_row_by_contains(tab.table, 0, "AREA-B")
    assert created_row >= 0

    tab.table.selectRow(created_row)
    qapp.processEvents()

    class FakeEditDialog:
        def __init__(self, *args, **kwargs):
            edited = kwargs["location"]
            self._location = edited

        def exec(self):
            return QDialog.DialogCode.Accepted

        @property
        def site_id(self):
            return self._location.site_id

        @property
        def location_code(self):
            return self._location.location_code

        @property
        def name(self):
            return "Area B Updated"

        @property
        def parent_location_id(self):
            return self._location.parent_location_id

        @property
        def location_type(self):
            return "ZONE"

        @property
        def criticality(self):
            return self._location.criticality.value

        @property
        def status(self):
            return self._location.status.value

        @property
        def description(self):
            return "Updated through the maintenance locations editor."

        @property
        def notes(self):
            return self._location.notes

        @property
        def is_active(self):
            return self._location.is_active

    monkeypatch.setattr(
        "ui.modules.maintenance_management.locations.tab.MaintenanceLocationEditDialog",
        FakeEditDialog,
    )
    tab.btn_edit_location.click()
    qapp.processEvents()
    created_row = _find_row_by_contains(tab.table, 0, "AREA-B")
    assert "Area B Updated" in tab.table.item(created_row, 1).text()

    tab.table.selectRow(created_row)
    qapp.processEvents()
    tab.btn_toggle_active.click()
    qapp.processEvents()
    assert _find_row_by_contains(tab.table, 0, "AREA-B") == -1
    tab.active_combo.setCurrentIndex(2)
    qapp.processEvents()
    created_row = _find_row_by_contains(tab.table, 0, "AREA-B")
    assert tab.table.item(created_row, 7).text() == "No"


def test_maintenance_systems_tab_supports_create_edit_and_toggle(qapp, services, monkeypatch):
    _enable_maintenance_module(services)
    _mute_message_boxes(monkeypatch)
    site, location, system, _asset, _symptom = _create_maintenance_context(services)

    tab = MaintenanceSystemsTab(
        system_service=services["maintenance_system_service"],
        location_service=services["maintenance_location_service"],
        site_service=services["site_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.table.rowCount() >= 1
    assert "PUMP-LINE" in tab.table.item(0, 0).text()

    class FakeCreateDialog:
        def __init__(self, *args, **kwargs):
            pass

        def exec(self):
            return QDialog.DialogCode.Accepted

        @property
        def site_id(self):
            return site.id

        @property
        def system_code(self):
            return "pack-line"

        @property
        def name(self):
            return "Packaging Line"

        @property
        def location_id(self):
            return location.id

        @property
        def parent_system_id(self):
            return system.id

        @property
        def system_type(self):
            return "PROCESS"

        @property
        def criticality(self):
            return "MEDIUM"

        @property
        def status(self):
            return "ACTIVE"

        @property
        def description(self):
            return "Secondary packaging system."

        @property
        def notes(self):
            return "Created from maintenance systems UI."

        @property
        def is_active(self):
            return True

    monkeypatch.setattr(
        "ui.modules.maintenance_management.systems.tab.MaintenanceSystemEditDialog",
        FakeCreateDialog,
    )
    tab.btn_new_system.click()
    qapp.processEvents()
    created_row = _find_row_by_contains(tab.table, 0, "PACK-LINE")
    assert created_row >= 0

    tab.table.selectRow(created_row)
    qapp.processEvents()

    class FakeEditDialog:
        def __init__(self, *args, **kwargs):
            edited = kwargs["system"]
            self._system = edited

        def exec(self):
            return QDialog.DialogCode.Accepted

        @property
        def site_id(self):
            return self._system.site_id

        @property
        def system_code(self):
            return self._system.system_code

        @property
        def name(self):
            return "Packaging Line Updated"

        @property
        def location_id(self):
            return self._system.location_id

        @property
        def parent_system_id(self):
            return self._system.parent_system_id

        @property
        def system_type(self):
            return "UTILITY"

        @property
        def criticality(self):
            return self._system.criticality.value

        @property
        def status(self):
            return self._system.status.value

        @property
        def description(self):
            return "Updated through the maintenance systems editor."

        @property
        def notes(self):
            return self._system.notes

        @property
        def is_active(self):
            return self._system.is_active

    monkeypatch.setattr(
        "ui.modules.maintenance_management.systems.tab.MaintenanceSystemEditDialog",
        FakeEditDialog,
    )
    tab.btn_edit_system.click()
    qapp.processEvents()
    created_row = _find_row_by_contains(tab.table, 0, "PACK-LINE")
    assert "Packaging Line Updated" in tab.table.item(created_row, 1).text()

    tab.table.selectRow(created_row)
    qapp.processEvents()
    tab.btn_toggle_active.click()
    qapp.processEvents()
    assert _find_row_by_contains(tab.table, 0, "PACK-LINE") == -1
    tab.active_combo.setCurrentIndex(2)
    qapp.processEvents()
    created_row = _find_row_by_contains(tab.table, 0, "PACK-LINE")
    assert tab.table.item(created_row, 7).text() == "No"


def test_maintenance_assets_tab_lists_assets_and_components(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, asset, _symptom = _create_maintenance_context(services)

    tab = MaintenanceAssetsTab(
        asset_service=services["maintenance_asset_service"],
        component_service=services["maintenance_asset_component_service"],
        site_service=services["site_service"],
        location_service=services["maintenance_location_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    assert tab.btn_filters.text().strip() == "Filters"
    assert tab.filter_panel.isHidden()
    tab.btn_filters.click()
    qapp.processEvents()
    assert not tab.filter_panel.isHidden()
    assert tab.btn_filters.text().strip() == "Hide Filters"
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.asset_table.rowCount() >= 1
    assert "Open Detail" in tab.selection_summary.text()
    row = _find_row_by_contains(tab.asset_table, 0, "PUMP-101")
    assert row >= 0
    tab.asset_table.selectRow(row)
    qapp.processEvents()
    assert "PUMP-101" in tab.selection_summary.text()
    assert tab.btn_open_detail.isEnabled()
    tab.btn_open_detail.click()
    qapp.processEvents()
    dialog = tab._detail_dialog
    assert dialog is not None
    assert dialog.title_label.text() == "PUMP-101 - Pump 101"
    assert dialog.component_table.rowCount() >= 1
    assert "Mechanical Seal" in dialog.component_table.item(0, 0).text()


def test_maintenance_requests_tab_lists_queue_and_linked_orders(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, _asset, _symptom = _create_maintenance_context(services)

    from ui.modules.maintenance_management import MaintenanceRequestsTab

    tab = MaintenanceRequestsTab(
        work_request_service=services["maintenance_work_request_service"],
        work_order_service=services["maintenance_work_order_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        location_service=services["maintenance_location_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.request_table.rowCount() >= 4
    assert tab.total_card._lbl_value.text() == str(tab.request_table.rowCount())
    assert "Open Detail" in tab.selection_summary.text()
    row = _find_row_by_contains(tab.request_table, 0, "WR-UI-OPEN")
    assert row >= 0
    tab.request_table.selectRow(row)
    qapp.processEvents()
    assert "WR-UI-OPEN" in tab.selection_summary.text()
    assert tab.btn_open_detail.isEnabled()
    tab.btn_open_detail.click()
    qapp.processEvents()
    dialog = tab._detail_dialog
    assert dialog is not None
    assert "Open seal leak" in dialog.title_label.text()
    assert dialog.linked_orders_table.rowCount() >= 1
    assert "WO-UI-OPEN" in dialog.linked_orders_table.item(0, 0).text()


def test_maintenance_work_orders_tab_lists_execution_queue_and_detail(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, _asset, _symptom = _create_maintenance_context(services)
    open_order = next(
        row
        for row in services["maintenance_work_order_service"].search_work_orders(search_text="WO-UI-OPEN")
        if row.work_order_code == "WO-UI-OPEN"
    )
    execution_doc = services["document_service"].create_document(
        document_code="DOC-UI-WO",
        title="Open Work Order Procedure",
        document_type="PROCEDURE",
        storage_kind="FILE_PATH",
        storage_uri="C:/docs/work-order-procedure.pdf",
    )
    services["maintenance_document_service"].link_existing_document(
        entity_type="work_order",
        entity_id=open_order.id,
        document_id=execution_doc.id,
        link_role="execution",
    )

    tab = MaintenanceWorkOrdersTab(
        work_order_service=services["maintenance_work_order_service"],
        work_order_task_service=services["maintenance_work_order_task_service"],
        work_order_task_step_service=services["maintenance_work_order_task_step_service"],
        material_requirement_service=services["maintenance_work_order_material_requirement_service"],
        labor_service=services["maintenance_labor_service"],
        document_service=services["maintenance_document_service"],
        work_request_service=services["maintenance_work_request_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        location_service=services["maintenance_location_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.work_order_table.rowCount() >= 3
    assert tab.total_card._lbl_value.text() == str(tab.work_order_table.rowCount())
    assert "Open Detail" in tab.selection_summary.text()
    row = _find_row_by_contains(tab.work_order_table, 0, "WO-UI-OPEN")
    assert row >= 0
    tab.work_order_table.selectRow(row)
    qapp.processEvents()
    assert "WO-UI-OPEN" in tab.selection_summary.text()
    assert tab.btn_open_detail.isEnabled()
    tab.btn_open_detail.click()
    qapp.processEvents()
    dialog = tab._detail_dialog
    assert dialog is not None
    assert "Open seal leak" in dialog.title_label.text()
    assert _top_level_labels(dialog.workbench.tree) == ["Overview", "Execution", "Tasks", "Task Steps", "Labor", "Materials", "Evidence"]
    assert dialog.task_table.rowCount() >= 1
    assert "Inspect seal housing" in dialog.task_table.item(0, 0).text()
    assert "Employee" in dialog.task_table.item(0, 1).text()
    assert dialog.step_table.rowCount() >= 1
    assert "Verify isolation" in dialog.step_table.item(0, 0).text()
    assert dialog.material_table.rowCount() >= 1
    assert "Seal kit" in dialog.material_table.item(0, 0).text()
    dialog.workbench.set_current_section("execution")
    qapp.processEvents()
    assert dialog.workbench.current_section_label.text() == "Execution"
    assert "Current focus" in dialog.execution_focus_summary.text()
    assert dialog.btn_start_quick_step.isEnabled()
    assert dialog.btn_done_quick_step.isEnabled()
    assert not dialog.btn_confirm_quick_step.isEnabled()
    dialog.btn_open_evidence.click()
    qapp.processEvents()
    assert dialog.workbench.current_section_key() == "evidence"
    dialog.workbench.set_current_section("evidence")
    qapp.processEvents()
    assert dialog.evidence_table.rowCount() >= 1
    assert "DOC-UI-WO" in dialog.evidence_table.item(0, 0).text()
    assert dialog.btn_preview_evidence.isEnabled()
    dialog.workbench.set_current_section("execution")
    qapp.processEvents()
    dialog.task_table.selectRow(0)
    dialog.step_table.selectRow(0)
    qapp.processEvents()
    assert dialog.btn_start_quick_step.isEnabled()
    assert dialog.btn_done_quick_step.isEnabled()
    assert not dialog.btn_confirm_quick_step.isEnabled()
    assert not dialog.btn_complete_quick_task.isEnabled()
    dialog.btn_start_quick_step.click()
    qapp.processEvents()
    assert dialog.workbench.current_section_key() == "execution"
    assert dialog.task_table.item(0, 2).text() == "In Progress"
    assert dialog.step_table.item(0, 1).text() == "In Progress"
    dialog.execution_measurement_edit.setText("81.5")
    assert dialog.step_measurement_edit.text() == "81.5"
    dialog.btn_done_quick_step.click()
    qapp.processEvents()
    assert dialog.step_table.item(0, 1).text() == "Done"
    assert "Measurement 81.5 C" in dialog.step_table.item(0, 3).text()
    assert dialog.btn_confirm_quick_step.isEnabled()
    dialog.btn_confirm_quick_step.click()
    qapp.processEvents()
    assert "Confirmed" in dialog.step_table.item(0, 3).text()
    dialog.workbench.set_current_section("labor")
    qapp.processEvents()
    dialog.labor_hours_spin.setValue(2.5)
    dialog.labor_note_edit.setText("Seal inspection labor")
    dialog.btn_add_labor.click()
    qapp.processEvents()
    assert dialog.labor_table.rowCount() >= 1
    assert dialog.labor_table.item(0, 1).text() == "2.50"
    assert "Seal inspection labor" in dialog.labor_table.item(0, 3).text()
    dialog.workbench.set_current_section("execution")
    qapp.processEvents()
    assert dialog.btn_complete_quick_task.isEnabled()
    dialog.btn_complete_quick_task.click()
    qapp.processEvents()
    assert dialog.workbench.current_section_key() == "execution"
    assert dialog.task_table.item(0, 2).text() == "Completed"
    assert "150" in dialog.task_table.item(0, 4).text()
    _select_combo_value(tab.responsibility_combo, "__EMPLOYEE__")
    qapp.processEvents()
    assert tab.work_order_table.rowCount() >= 1
    assert "WO-UI-OPEN" in tab.work_order_table.item(0, 0).text()


def test_maintenance_work_order_detail_supports_evidence_capture_link_preview_and_unlink(
    qapp,
    services,
    monkeypatch,
):
    _enable_maintenance_module(services)
    _mute_message_boxes(monkeypatch)
    site, _location, _system, _asset, _symptom = _create_maintenance_context(services)
    open_order = next(
        row
        for row in services["maintenance_work_order_service"].search_work_orders(search_text="WO-UI-OPEN")
        if row.work_order_code == "WO-UI-OPEN"
    )
    structure = services["document_service"].create_document_structure(
        structure_code="WO_EVIDENCE",
        name="Work Order Evidence",
        object_scope="WORK_ORDER",
        default_document_type="GENERAL",
    )
    reusable_document = services["document_service"].create_document(
        document_code="DOC-UI-LINK",
        title="Reusable Execution Checklist",
        document_type="GENERAL",
        document_structure_id=structure.id,
        storage_kind="FILE_PATH",
        storage_uri="C:/docs/reusable-execution-checklist.pdf",
    )

    class FakeCaptureDialog:
        def __init__(self, *args, **kwargs):
            self.attachments = ["C:/evidence/seal-inspection-photo.pdf"]
            self.document_type = "GENERAL"
            self.document_structure_id = structure.id
            self.business_version_label = "REV-A"
            self.source_system = "maintenance_execution"
            self.link_role = "evidence"
            self.notes = "Captured during execution"

        def exec(self):
            return QDialog.DialogCode.Accepted

    class FakeLinkDialog:
        def __init__(self, *args, **kwargs):
            self.document_id = reusable_document.id
            self.link_role = "reference"

        def exec(self):
            return QDialog.DialogCode.Accepted

    preview_calls: list[str] = []
    link_counts: list[int] = []

    class FakePreviewDialog:
        def __init__(self, *, document, parent=None):
            self._document = document

        def exec(self):
            preview_calls.append(self._document.document_code)
            return QDialog.DialogCode.Accepted

    class FakeLinksDialog:
        def __init__(self, *, document, links, parent=None):
            self._document = document
            self._links = links

        def exec(self):
            link_counts.append(len(self._links))
            return QDialog.DialogCode.Accepted

    monkeypatch.setattr(
        "ui.modules.maintenance_management.work_orders.dialogs.MaintenanceWorkOrderEvidenceCaptureDialog",
        FakeCaptureDialog,
    )
    monkeypatch.setattr(
        "ui.modules.maintenance_management.work_orders.dialogs.MaintenanceWorkOrderEvidenceLinkDialog",
        FakeLinkDialog,
    )
    monkeypatch.setattr(
        "ui.modules.maintenance_management.work_orders.dialogs.DocumentPreviewDialog",
        FakePreviewDialog,
    )
    monkeypatch.setattr(
        "ui.modules.maintenance_management.work_orders.dialogs.DocumentLinksDialog",
        FakeLinksDialog,
    )

    tab = MaintenanceWorkOrdersTab(
        work_order_service=services["maintenance_work_order_service"],
        work_order_task_service=services["maintenance_work_order_task_service"],
        work_order_task_step_service=services["maintenance_work_order_task_step_service"],
        material_requirement_service=services["maintenance_work_order_material_requirement_service"],
        labor_service=services["maintenance_labor_service"],
        document_service=services["maintenance_document_service"],
        work_request_service=services["maintenance_work_request_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        location_service=services["maintenance_location_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()
    row = _find_row_by_contains(tab.work_order_table, 0, "WO-UI-OPEN")
    assert row >= 0
    tab.work_order_table.selectRow(row)
    qapp.processEvents()
    tab.btn_open_detail.click()
    qapp.processEvents()
    dialog = tab._detail_dialog
    assert dialog is not None

    dialog.workbench.set_current_section("evidence")
    qapp.processEvents()
    initial_count = dialog.evidence_table.rowCount()
    dialog.btn_capture_evidence.click()
    qapp.processEvents()
    assert dialog.workbench.current_section_key() == "evidence"
    assert dialog.evidence_table.rowCount() == initial_count + 1
    assert any(
        "WO_EVIDENCE" in dialog.evidence_table.item(row_index, 3).text()
        for row_index in range(dialog.evidence_table.rowCount())
    )

    dialog.btn_link_evidence.click()
    qapp.processEvents()
    linked_row = _find_row_by_contains(dialog.evidence_table, 0, "DOC-UI-LINK")
    assert linked_row >= 0
    dialog.evidence_table.selectRow(linked_row)
    qapp.processEvents()
    dialog.btn_preview_evidence.click()
    dialog.btn_view_evidence_links.click()
    qapp.processEvents()
    assert preview_calls == ["DOC-UI-LINK"]
    assert link_counts and link_counts[-1] >= 1

    before_unlink = dialog.evidence_table.rowCount()
    dialog.btn_unlink_evidence.click()
    qapp.processEvents()
    assert dialog.evidence_table.rowCount() == before_unlink - 1
    assert _find_row_by_contains(dialog.evidence_table, 0, "DOC-UI-LINK") == -1


def test_maintenance_documents_tab_lists_linked_documents(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, asset, _symptom = _create_maintenance_context(services)
    document = services["document_service"].create_document(
        document_code="DOC-UI-MNT",
        title="Maintenance SOP",
        document_type="PROCEDURE",
        storage_kind="FILE_PATH",
        storage_uri="C:/docs/maintenance-sop.pdf",
    )
    services["maintenance_document_service"].link_existing_document(
        entity_type="asset",
        entity_id=asset.id,
        document_id=document.id,
        link_role="reference",
    )

    tab = MaintenanceDocumentsTab(
        document_service=services["maintenance_document_service"],
        site_service=services["site_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.document_table.rowCount() >= 1
    row = _find_row_by_contains(tab.document_table, 0, "DOC-UI-MNT")
    assert row >= 0
    tab.document_table.selectRow(row)
    qapp.processEvents()
    assert tab.detail_title.text() == "Maintenance SOP"
    assert "Asset: PUMP-101 - Pump 101" in tab.metadata_labels["linked_record"].text()
    assert tab.metadata_labels["type"].text() == "Procedure"


def test_maintenance_sensors_tab_lists_sensor_register_and_exception_queue(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, asset, _symptom = _create_maintenance_context(services)

    tab = MaintenanceSensorsTab(
        sensor_service=services["maintenance_sensor_service"],
        sensor_reading_service=services["maintenance_sensor_reading_service"],
        sensor_source_mapping_service=services["maintenance_sensor_source_mapping_service"],
        sensor_exception_service=services["maintenance_sensor_exception_service"],
        integration_source_service=services["maintenance_integration_source_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        component_service=services["maintenance_asset_component_service"],
        location_service=services["maintenance_location_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    assert tab.btn_filters.text().strip() == "Filters"
    assert tab.filter_panel.isHidden()
    tab.btn_filters.click()
    qapp.processEvents()
    assert not tab.filter_panel.isHidden()
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()
    _select_combo_value(tab.asset_combo, asset.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.workbench.current_section_key() == "sensor_register"
    assert tab.workbench.current_section_label.text() == "Sensor Register"
    assert _top_level_labels(tab.workbench.tree) == [
        "Sensor Register",
        "Integration Queue",
        "Exception Queue",
    ]
    assert not tab.workbench.tree_panel.isHidden()
    tab.workbench.btn_toggle_tree.click()
    qapp.processEvents()
    assert tab.workbench.tree_panel.isHidden()
    assert tab.workbench.btn_toggle_tree.text() == ">> Show Sections"
    tab.workbench.btn_toggle_tree.click()
    qapp.processEvents()
    assert not tab.workbench.tree_panel.isHidden()
    assert tab.sensor_table.rowCount() >= 1
    assert "RUN-HOURS-101" in tab.sensor_table.item(0, 0).text()
    assert tab.btn_view_sensor_detail.isEnabled()
    tab.btn_view_sensor_detail.click()
    qapp.processEvents()
    dialog = tab._sensor_detail_dialog
    assert dialog is not None
    assert "RUN-HOURS-101" in dialog.title_label.text()
    assert _top_level_labels(dialog.workbench.tree) == ["Overview", "Recent Readings", "Source Mappings"]
    dialog.workbench.set_current_section("readings")
    qapp.processEvents()
    assert dialog.workbench.current_section_label.text() == "Recent Readings"
    assert dialog.readings_table.rowCount() >= 1
    dialog.workbench.set_current_section("mappings")
    qapp.processEvents()
    assert dialog.mapping_table.rowCount() >= 1
    tab.workbench.set_current_section("integration_queue")
    qapp.processEvents()
    assert tab.integration_table.rowCount() >= 1
    tab.workbench.set_current_section("exception_queue")
    qapp.processEvents()
    assert tab.exception_table.rowCount() >= 1
    assert tab.attention_card._lbl_value.text() != "0"
    assert tab.open_exceptions_card._lbl_value.text() != "0"


def test_maintenance_planner_tab_surfaces_backlog_material_and_recurring_queues(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, asset, _symptom = _create_maintenance_context(services)

    tab = MaintenancePlannerTab(
        work_request_service=services["maintenance_work_request_service"],
        work_order_service=services["maintenance_work_order_service"],
        material_requirement_service=services["maintenance_work_order_material_requirement_service"],
        preventive_plan_service=services["maintenance_preventive_plan_service"],
        preventive_generation_service=services["maintenance_preventive_generation_service"],
        reliability_service=services["maintenance_reliability_service"],
        sensor_exception_service=services["maintenance_sensor_exception_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    assert tab.btn_filters.text().strip() == "Filters"
    assert tab.filter_panel.isHidden()
    tab.btn_filters.click()
    qapp.processEvents()
    assert not tab.filter_panel.isHidden()
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()
    _select_combo_value(tab.asset_combo, asset.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.workbench.current_section_key() == "request_intake"
    assert _top_level_labels(tab.workbench.tree) == [
        "Request Intake",
        "Backlog and Scheduling",
        "Material Readiness",
        "Preventive Readiness",
        "Recurring Failure Review",
    ]
    tab.workbench.btn_toggle_tree.click()
    qapp.processEvents()
    assert tab.workbench.tree_panel.isHidden()
    tab.workbench.btn_toggle_tree.click()
    qapp.processEvents()
    assert not tab.workbench.tree_panel.isHidden()
    assert tab.request_table.rowCount() >= 1
    tab.workbench.set_current_section("backlog")
    qapp.processEvents()
    assert tab.work_order_table.rowCount() >= 1
    tab.workbench.set_current_section("preventive_readiness")
    qapp.processEvents()
    assert tab.preventive_table.rowCount() >= 1
    tab.workbench.set_current_section("material_readiness")
    qapp.processEvents()
    assert tab.material_table.rowCount() >= 1
    tab.workbench.set_current_section("recurring_failure_review")
    qapp.processEvents()
    assert tab.recurring_table.rowCount() >= 1
    assert tab.backlog_card._lbl_value.text() != "0"
    assert tab.preventive_card._lbl_value.text() != "0"
    assert tab.material_card._lbl_value.text() != "0"
    assert "WO-UI-OPEN" in tab.work_order_table.item(0, 0).text()
    due_row = _find_row_by_contains(tab.preventive_table, 0, "PM-UI-DUE")
    blocked_row = _find_row_by_contains(tab.preventive_table, 0, "PM-UI-BLOCKED")
    assert due_row >= 0
    assert blocked_row >= 0
    assert due_row < blocked_row
    assert tab.preventive_table.item(due_row, 2).text() == "Due"
    assert tab.preventive_table.item(blocked_row, 2).text() == "Blocked"


def test_maintenance_preventive_tab_surfaces_due_blocked_and_inactive_plan_states(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, _asset, _symptom = _create_maintenance_context(services)

    tab = MaintenancePreventivePlansTab(
        preventive_plan_service=services["maintenance_preventive_plan_service"],
        preventive_plan_task_service=services["maintenance_preventive_plan_task_service"],
        preventive_generation_service=services["maintenance_preventive_generation_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        system_service=services["maintenance_system_service"],
        sensor_service=services["maintenance_sensor_service"],
        task_template_service=services["maintenance_task_template_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    assert tab.btn_filters.text().strip() == "Filters"
    assert tab.filter_panel.isHidden()
    tab.btn_filters.click()
    qapp.processEvents()
    assert not tab.filter_panel.isHidden()
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.plan_table.rowCount() >= 4
    assert tab.due_now_card._lbl_value.text() != "0"
    assert tab.due_soon_card._lbl_value.text() != "0"
    assert tab.blocked_card._lbl_value.text() != "0"
    assert "Open Detail" in tab.selection_summary.text()

    row = _find_row_by_contains(tab.plan_table, 0, "PM-UI-DUE")
    assert row >= 0
    tab.plan_table.selectRow(row)
    qapp.processEvents()
    assert "PM-UI-DUE" in tab.selection_summary.text()
    assert tab.btn_open_detail.isEnabled()
    tab.btn_open_detail.click()
    qapp.processEvents()
    dialog = tab._detail_dialog
    assert dialog is not None
    assert "Due seal inspection" in dialog.title_label.text()
    assert "Due state: Due" in dialog.overview_summary.text()
    assert dialog.task_table.rowCount() >= 1
    assert "PM-SEAL-CHECK" in dialog.task_table.item(0, 1).text()

    _select_combo_value(tab.due_state_combo, "BLOCKED")
    qapp.processEvents()
    assert tab.plan_table.rowCount() >= 1
    assert "PM-UI-BLOCKED" in tab.plan_table.item(0, 0).text()


def test_maintenance_dashboard_tab_surfaces_reliability_metrics(qapp, services):
    _enable_maintenance_module(services)
    site, _location, _system, asset, _symptom = _create_maintenance_context(services)

    tab = MaintenanceDashboardTab(
        reliability_service=services["maintenance_reliability_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        location_service=services["maintenance_location_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()
    _select_combo_value(tab.asset_combo, asset.id)
    qapp.processEvents()

    assert tab.context_badge.text() == "Context: Default Organization"
    assert tab.window_badge.text() == "90 day window"
    assert tab.root_cause_table.rowCount() >= 1
    assert tab.recurring_table.rowCount() >= 1
    assert "Pump 101" in tab.recurring_table.item(0, 0).text()
    assert "Misalignment" in tab.root_cause_table.item(0, 1).text()


def test_maintenance_reliability_tab_exports_report_packs(qapp, services, tmp_path, monkeypatch):
    _enable_maintenance_module(services)
    _mute_message_boxes(monkeypatch)
    site, _location, _system, asset, symptom = _create_maintenance_context(services)

    tab = MaintenanceReliabilityTab(
        reliability_service=services["maintenance_reliability_service"],
        reporting_service=services["maintenance_reporting_service"],
        failure_code_service=services["maintenance_failure_code_service"],
        site_service=services["site_service"],
        asset_service=services["maintenance_asset_service"],
        location_service=services["maintenance_location_service"],
        system_service=services["maintenance_system_service"],
        platform_runtime_application_service=services["platform_runtime_application_service"],
        user_session=services["user_session"],
    )
    assert tab.layout().itemAt(1).widget() is tab.export_surface
    assert tab.layout().itemAt(2).widget() is tab.control_surface
    assert tab.btn_filters.text().strip() == "Filters"
    assert tab.filter_panel.isHidden()
    assert tab.workbench.current_section_key() == "suggestions"
    assert tab.workbench.current_section_label.text() == "Suggestions"
    assert _top_level_labels(tab.workbench.tree) == [
        "Suggestions",
        "Root Causes",
        "Recurring Failures",
    ]
    tab.workbench.btn_toggle_tree.click()
    qapp.processEvents()
    assert tab.workbench.tree_panel.isHidden()
    tab.workbench.btn_toggle_tree.click()
    qapp.processEvents()
    assert not tab.workbench.tree_panel.isHidden()
    _select_combo_value(tab.site_combo, site.id)
    qapp.processEvents()
    _select_combo_value(tab.asset_combo, asset.id)
    qapp.processEvents()
    _select_combo_value(tab.failure_code_combo, symptom.failure_code)
    qapp.processEvents()

    backlog_path = tab.export_backlog_excel(tmp_path / "maintenance-ui-backlog.xlsx")
    recurring_path = tab.export_recurring_excel(tmp_path / "maintenance-ui-recurring.xlsx")
    exception_path = tab.export_exception_excel(tmp_path / "maintenance-ui-exceptions.xlsx")
    downtime_path = tab.export_downtime_pdf(tmp_path / "maintenance-ui-downtime.pdf")

    assert tab.export_badge.text() == "Export Enabled"
    assert tab.suggestions_table.rowCount() >= 1
    tab.workbench.set_current_section("root_causes")
    qapp.processEvents()
    assert tab.root_cause_table.rowCount() >= 1
    assert tab.workbench.current_section_label.text() == "Root Causes"
    tab.workbench.set_current_section("recurring_failures")
    qapp.processEvents()
    assert tab.recurring_table.rowCount() >= 1
    assert tab.workbench.current_section_label.text() == "Recurring Failures"
    assert backlog_path is not None
    assert recurring_path is not None
    assert exception_path is not None
    assert downtime_path is not None
    assert Path(backlog_path).exists()
    assert Path(recurring_path).exists()
    assert Path(exception_path).exists()
    assert Path(downtime_path).exists()
    workbook = load_workbook(backlog_path)
    recurring_workbook = load_workbook(recurring_path)
    exception_workbook = load_workbook(exception_path)
    assert {"Summary", "Open Work Orders", "Top Root Causes", "Recurring Failure Patterns"} <= set(workbook.sheetnames)
    assert {"Summary", "Leading Root Causes", "Recurring Failure Review"} <= set(recurring_workbook.sheetnames)
    assert {"Summary", "Open Sensor Exceptions", "Recurring Failure Signals"} <= set(exception_workbook.sheetnames)


def test_main_window_exposes_maintenance_workspaces_when_module_is_enabled(
    qapp,
    services,
    repo_workspace,
    monkeypatch,
):
    _enable_maintenance_module(services)
    _create_maintenance_context(services)
    store = make_settings_store(repo_workspace, prefix="maintenance-shell")
    monkeypatch.setattr("ui.platform.shell.main_window.MainWindowSettingsStore", lambda: store)
    monkeypatch.setattr(MainWindow, "_run_startup_update_check", lambda self: None)

    window = MainWindow(services)
    labels = [window.tabs.tabText(i) for i in range(window.tabs.count())]

    assert "Maintenance Dashboard" in labels
    assert "Locations" in labels
    assert "Systems" in labels
    assert "Assets" in labels
    assert "Planner" in labels
    assert "Preventive Plans" in labels
    assert "Sensors" in labels
    assert "Requests" in labels
    assert "Work Orders" in labels
    assert "Documents" in labels
    assert "Reliability" in labels

    maintenance_section = window.shell_navigation.tree.topLevelItem(3)
    assert maintenance_section.text(0) == "Maintenance Management"
    assert _child_labels(maintenance_section) == ["Overview", "Libraries", "Records", "Planning", "Analytics"]
    assert _child_labels(maintenance_section.child(0)) == ["Maintenance Dashboard"]
    assert _child_labels(maintenance_section.child(1)) == ["Locations", "Systems"]
    assert _child_labels(maintenance_section.child(2)) == ["Assets", "Sensors", "Requests", "Work Orders", "Documents"]
    assert _child_labels(maintenance_section.child(3)) == ["Preventive Plans", "Planner"]
    assert _child_labels(maintenance_section.child(4)) == ["Reliability"]
