from __future__ import annotations

from datetime import date
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from src.core.platform.auth.domain.session import UserSessionPrincipal
from core.platform.common.exceptions import BusinessRuleError
from src.core.platform.party.domain import PartyType
from tests.ui_runtime_helpers import login_as


def _suffix() -> str:
    return uuid4().hex[:6].upper()


def _enable_inventory_module(services) -> None:
    services["module_catalog_service"].set_module_state(
        "inventory_procurement",
        licensed=True,
        enabled=True,
    )


def _create_inventory_context(services):
    suffix = _suffix()
    site = services["site_service"].create_site(
        site_code=f"INV-{suffix}",
        name=f"Inventory Site {suffix}",
        currency_code="EUR",
    )
    supplier = services["party_service"].create_party(
        party_code=f"SUP-{suffix}",
        party_name=f"Inventory Supplier {suffix}",
        party_type=PartyType.SUPPLIER,
    )
    storeroom = services["inventory_service"].create_storeroom(
        storeroom_code=f"ST-{suffix}",
        name=f"Storeroom {suffix}",
        site_id=site.id,
        status="ACTIVE",
        manager_party_id=supplier.id,
    )
    item = services["inventory_item_service"].create_item(
        item_code=f"ITEM-{suffix}",
        name=f"Item {suffix}",
        status="ACTIVE",
        stock_uom="EA",
        preferred_party_id=supplier.id,
        reorder_point=2,
        reorder_qty=4,
        max_qty=10,
    )
    return site, supplier, storeroom, item


def _create_procurement_context(services):
    auth = services["auth_service"]
    auth.register_user("inventory-report-buyer", "StrongPass123", role_names=["inventory_manager"])
    auth.register_user("inventory-report-approver", "StrongPass123", role_names=["approver"])
    _enable_inventory_module(services)

    site, supplier, storeroom, item = _create_inventory_context(services)
    procurement = services["inventory_procurement_service"]
    purchasing = services["inventory_purchasing_service"]
    approvals = services["approval_service"]

    login_as(services, "inventory-report-buyer", "StrongPass123")
    requisition = procurement.create_requisition(
        requesting_site_id=site.id,
        requesting_storeroom_id=storeroom.id,
        purpose="Restock rotating spare",
        needed_by_date=date(2026, 4, 12),
    )
    requisition_line = procurement.add_requisition_line(
        requisition.id,
        stock_item_id=item.id,
        quantity_requested=5,
        estimated_unit_cost=22.5,
        suggested_supplier_party_id=supplier.id,
    )
    requisition = procurement.submit_requisition(requisition.id)

    login_as(services, "inventory-report-approver", "StrongPass123")
    approvals.approve_and_apply(requisition.approval_request_id, note="Approved for procurement reporting")

    login_as(services, "inventory-report-buyer", "StrongPass123")
    purchase_order = purchasing.create_purchase_order(
        site_id=site.id,
        supplier_party_id=supplier.id,
        currency_code="EUR",
        source_requisition_id=requisition.id,
        expected_delivery_date=date(2026, 4, 18),
    )
    purchase_order_line = purchasing.add_purchase_order_line(
        purchase_order.id,
        stock_item_id=item.id,
        destination_storeroom_id=storeroom.id,
        quantity_ordered=5,
        unit_price=20.0,
        source_requisition_line_id=requisition_line.id,
    )
    purchase_order = purchasing.submit_purchase_order(purchase_order.id)

    login_as(services, "inventory-report-approver", "StrongPass123")
    approvals.approve_and_apply(purchase_order.approval_request_id, note="Approved PO for reporting")

    login_as(services, "inventory-report-buyer", "StrongPass123")
    receipt = purchasing.post_receipt(
        purchase_order.id,
        receipt_lines=[
            {
                "purchase_order_line_id": purchase_order_line.id,
                "quantity_accepted": 4,
                "quantity_rejected": 1,
                "unit_cost": 20.0,
            }
        ],
        supplier_delivery_reference=f"DEL-{_suffix()}",
    )
    return site, supplier, storeroom, item, requisition, purchase_order, receipt


def test_inventory_data_exchange_service_imports_and_exports_items_and_storerooms(services, tmp_path):
    auth = services["auth_service"]
    auth.register_user("inventory-import-user", "StrongPass123", role_names=["inventory_manager"])
    _enable_inventory_module(services)

    service = services["inventory_data_exchange_service"]
    site = services["site_service"].create_site(site_code="IMP-HUB", name="Import Hub", currency_code="EUR")
    supplier = services["party_service"].create_party(
        party_code="IMP-SUP",
        party_name="Import Supplier",
        party_type=PartyType.SUPPLIER,
    )
    login_as(services, "inventory-import-user", "StrongPass123")

    storerooms_csv = tmp_path / "storerooms.csv"
    storerooms_csv.write_text(
        "storeroom_code,name,site_code,status,manager_party_code,allows_receiving\n"
        "IMP-MAIN,Import Main,IMP-HUB,ACTIVE,IMP-SUP,true\n",
        encoding="utf-8",
    )
    storeroom_preview = service.preview_csv("storerooms", storerooms_csv)
    storeroom_summary = service.import_csv("storerooms", storerooms_csv)

    items_csv = tmp_path / "items.csv"
    items_csv.write_text(
        "item_code,name,status,stock_uom,preferred_party_code,reorder_point,reorder_qty,max_qty\n"
        "IMP-ITEM,Imported Item,ACTIVE,EA,IMP-SUP,3,6,12\n",
        encoding="utf-8",
    )
    item_preview = service.preview_csv("items", items_csv)
    item_summary = service.import_csv("items", items_csv)

    items_export = service.export_csv("items", tmp_path / "items-export.csv", active_only=True)
    storerooms_export = service.export_csv("storerooms", tmp_path / "storerooms-export.csv", active_only=True)

    assert storeroom_preview.created_count == 1
    assert storeroom_summary.created_count == 1
    assert item_preview.created_count == 1
    assert item_summary.created_count == 1
    assert "IMP-ITEM" in items_export.file_path.read_text(encoding="utf-8")
    assert supplier.party_code in items_export.file_path.read_text(encoding="utf-8")
    assert "IMP-MAIN" in storerooms_export.file_path.read_text(encoding="utf-8")
    assert site.site_code in storerooms_export.file_path.read_text(encoding="utf-8")


def test_inventory_data_exchange_service_exports_requisitions_purchase_orders_and_receipts(services, tmp_path):
    (
        _site,
        supplier,
        _storeroom,
        _item,
        requisition,
        purchase_order,
        receipt,
    ) = _create_procurement_context(services)
    service = services["inventory_data_exchange_service"]

    requisition_export = service.export_csv("requisitions", tmp_path / "requisitions.csv")
    purchase_order_export = service.export_csv("purchase_orders", tmp_path / "purchase-orders.csv")
    receipt_export = service.export_csv("receipts", tmp_path / "receipts.csv")

    requisition_text = requisition_export.file_path.read_text(encoding="utf-8")
    purchase_order_text = purchase_order_export.file_path.read_text(encoding="utf-8")
    receipt_text = receipt_export.file_path.read_text(encoding="utf-8")

    assert requisition.requisition_number in requisition_text
    assert purchase_order.po_number in purchase_order_text
    assert receipt.receipt_number in receipt_text
    assert supplier.party_code in purchase_order_text


def test_inventory_reporting_service_generates_stock_and_procurement_reports(services, tmp_path):
    site, supplier, storeroom, item, requisition, purchase_order, receipt = _create_procurement_context(services)
    reporting = services["inventory_reporting_service"]

    stock_csv = reporting.generate_stock_status_csv(tmp_path / "stock-status.csv", site_id=site.id)
    stock_excel = reporting.generate_stock_status_excel(tmp_path / "stock-status.xlsx", storeroom_id=storeroom.id)
    procurement_csv = reporting.generate_procurement_overview_csv(
        tmp_path / "procurement.csv",
        site_id=site.id,
        supplier_party_id=supplier.id,
    )
    procurement_excel = reporting.generate_procurement_overview_excel(
        tmp_path / "procurement.xlsx",
        site_id=site.id,
        supplier_party_id=supplier.id,
    )

    stock_text = stock_csv.file_path.read_text(encoding="utf-8")
    procurement_text = procurement_csv.file_path.read_text(encoding="utf-8")
    stock_workbook = load_workbook(stock_excel.file_path)
    procurement_workbook = load_workbook(procurement_excel.file_path)

    assert item.item_code in stock_text
    assert "Inventory Stock Status" in stock_text
    assert requisition.requisition_number in procurement_text
    assert purchase_order.po_number in procurement_text
    assert receipt.receipt_number in procurement_text
    assert {"Summary", "Balances"} <= set(stock_workbook.sheetnames)
    assert {"Summary", "Requisitions", "Purchase Orders", "Receipts"} <= set(procurement_workbook.sheetnames)


def test_inventory_import_and_reporting_services_require_runtime_permissions(services, tmp_path):
    _enable_inventory_module(services)
    services["user_session"].set_principal(
        UserSessionPrincipal(
            user_id="u-inventory-read",
            username="inventory-read-only",
            display_name="Inventory Read Only",
            role_names=frozenset({"viewer"}),
            permissions=frozenset({"inventory.read", "report.view"}),
        )
    )

    with pytest.raises(BusinessRuleError, match="Permission denied") as import_exc:
        services["inventory_data_exchange_service"].get_import_schema("items")
    with pytest.raises(BusinessRuleError, match="Permission denied") as report_exc:
        services["inventory_reporting_service"].generate_stock_status_csv(tmp_path / "denied.csv")

    assert import_exc.value.code == "PERMISSION_DENIED"
    assert report_exc.value.code == "PERMISSION_DENIED"


def test_inventory_data_exchange_imports_procurement_documents_and_records_runtime_executions(services, tmp_path):
    _enable_inventory_module(services)
    login_as(services, "admin", "ChangeMe123!")
    services["auth_service"].register_user("inventory-import-approver", "StrongPass123", role_names=["approver"])

    site = services["site_service"].create_site(
        site_code="IMP-PROC-SITE",
        name="Import Procurement Site",
        currency_code="EUR",
    )
    supplier = services["party_service"].create_party(
        party_code="IMP-PROC-SUP",
        party_name="Import Procurement Supplier",
        party_type=PartyType.SUPPLIER,
    )
    services["inventory_service"].create_storeroom(
        storeroom_code="IMP-PROC-ST",
        name="Import Procurement Stores",
        site_id=site.id,
        status="ACTIVE",
        requires_supplier_reference_for_receipt=True,
    )
    services["inventory_item_service"].create_item(
        item_code="IMP-PROC-ITEM",
        name="Import Procurement Item",
        status="ACTIVE",
        stock_uom="EA",
        is_lot_tracked=True,
        shelf_life_days=30,
        preferred_party_id=supplier.id,
    )
    exchange = services["inventory_data_exchange_service"]

    requisitions_csv = tmp_path / "requisitions-import.csv"
    requisitions_csv.write_text(
        "requisition_number,requesting_site_code,requesting_storeroom_code,purpose,needed_by_date,priority,source_reference_type,source_reference_id,status,line_number,item_code,quantity_requested,uom,estimated_unit_cost,suggested_supplier_code\n"
        "REQ-IMPORT-001,IMP-PROC-SITE,IMP-PROC-ST,Imported maintenance demand,2026-05-01,HIGH,maintenance_work_order,MWO-001,DRAFT,1,IMP-PROC-ITEM,5,EA,12.5,IMP-PROC-SUP\n",
        encoding="utf-8",
    )
    requisition_summary = exchange.import_csv("requisitions", requisitions_csv)
    requisition = services["inventory_procurement_service"].find_requisition_by_number("REQ-IMPORT-001")
    requisition = services["inventory_procurement_service"].submit_requisition(requisition.id)

    login_as(services, "inventory-import-approver", "StrongPass123")
    services["approval_service"].approve_and_apply(requisition.approval_request_id, note="Approve imported requisition")
    login_as(services, "admin", "ChangeMe123!")
    requisition = services["inventory_procurement_service"].find_requisition_by_number("REQ-IMPORT-001")

    purchase_orders_csv = tmp_path / "purchase-orders-import.csv"
    purchase_orders_csv.write_text(
        "po_number,site_code,supplier_code,currency_code,source_requisition_number,expected_delivery_date,supplier_reference,status,line_number,item_code,destination_storeroom_code,quantity_ordered,uom,unit_price,source_requisition_line_ref\n"
        "PO-IMPORT-001,IMP-PROC-SITE,IMP-PROC-SUP,EUR,REQ-IMPORT-001,2026-05-10,SUP-REF-001,DRAFT,1,IMP-PROC-ITEM,IMP-PROC-ST,5,EA,11.0,REQ-IMPORT-001:1\n",
        encoding="utf-8",
    )
    po_summary = exchange.import_csv("purchase_orders", purchase_orders_csv)
    purchase_order = services["inventory_purchasing_service"].find_purchase_order_by_number("PO-IMPORT-001")
    purchase_order = services["inventory_purchasing_service"].submit_purchase_order(purchase_order.id)

    login_as(services, "inventory-import-approver", "StrongPass123")
    services["approval_service"].approve_and_apply(purchase_order.approval_request_id, note="Approve imported PO")
    login_as(services, "admin", "ChangeMe123!")
    purchase_order = services["inventory_purchasing_service"].find_purchase_order_by_number("PO-IMPORT-001")
    purchase_order = services["inventory_purchasing_service"].send_purchase_order(purchase_order.id)

    receipts_csv = tmp_path / "receipts-import.csv"
    receipts_csv.write_text(
        "receipt_number,purchase_order_number,receipt_date,supplier_delivery_reference,line_number,purchase_order_line_number,quantity_accepted,quantity_rejected,unit_cost,lot_number,expiry_date\n"
        "RCV-IMPORT-001,PO-IMPORT-001,2026-05-10,DEL-IMPORT-001,1,1,5,0,11.0,LOT-IMPORT-001,2026-06-15\n",
        encoding="utf-8",
    )
    receipt_summary = exchange.import_csv("receipts", receipts_csv)
    receipt = services["inventory_purchasing_service"].find_receipt_by_number("RCV-IMPORT-001")
    receipt_lines = services["inventory_purchasing_service"].list_receipt_lines(receipt.id)
    executions = services["runtime_execution_service"].list_recent(limit=10)

    assert requisition_summary.created_count == 1
    assert po_summary.created_count == 1
    assert receipt_summary.created_count == 1
    assert requisition is not None
    assert requisition.status.value == "APPROVED"
    assert purchase_order is not None
    assert purchase_order.status.value == "SENT"
    assert receipt is not None
    assert receipt_lines[0].lot_number == "LOT-IMPORT-001"
    assert str(receipt_lines[0].expiry_date) == "2026-06-15"
    assert {(execution.operation_type, execution.operation_key, execution.status) for execution in executions} >= {
        ("import", "requisitions", "COMPLETED"),
        ("import", "purchase_orders", "COMPLETED"),
        ("import", "receipts", "COMPLETED"),
    }
