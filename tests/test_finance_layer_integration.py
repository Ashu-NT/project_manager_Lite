from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from core.models import CostType
from core.reporting import api as reporting_api


def _seed_finance_project(services) -> str:
    ps = services["project_service"]
    ts = services["task_service"]
    rs = services["resource_service"]
    prs = services["project_resource_service"]
    cs = services["cost_service"]

    project = ps.create_project(
        "Finance Integration",
        start_date=date(2024, 1, 8),
        end_date=date(2024, 2, 29),
        planned_budget=5000.0,
        currency="EUR",
    )
    pid = project.id

    task_a = ts.create_task(pid, "Execution A", start_date=date(2024, 1, 8), duration_days=5)
    ts.create_task(pid, "Execution B", start_date=date(2024, 1, 15), duration_days=3)

    resource = rs.create_resource("Engineer A", "Developer", hourly_rate=60.0, currency_code="EUR")
    pr = prs.add_to_project(
        project_id=pid,
        resource_id=resource.id,
        planned_hours=20.0,
        hourly_rate=60.0,
        currency_code="EUR",
    )
    assignment = ts.assign_project_resource(
        task_id=task_a.id,
        project_resource_id=pr.id,
        allocation_percent=100.0,
    )
    ts.set_assignment_hours(assignment.id, 6.0)

    cs.add_cost_item(
        project_id=pid,
        task_id=task_a.id,
        description="Materials",
        planned_amount=400.0,
        committed_amount=300.0,
        actual_amount=250.0,
        cost_type=CostType.MATERIAL,
        incurred_date=date(2024, 1, 20),
        currency_code="EUR",
    )
    cs.add_cost_item(
        project_id=pid,
        description="Office Overhead",
        planned_amount=200.0,
        committed_amount=100.0,
        actual_amount=80.0,
        cost_type=CostType.OVERHEAD,
        incurred_date=date(2024, 1, 28),
        currency_code="EUR",
    )
    cs.add_cost_item(
        project_id=pid,
        task_id=task_a.id,
        description="Manual labor adjustment",
        planned_amount=150.0,
        committed_amount=75.0,
        actual_amount=50.0,
        cost_type=CostType.LABOR,
        incurred_date=date(2024, 1, 25),
        currency_code="EUR",
    )
    return pid


def test_finance_snapshot_builds_policy_aligned_ledger_cashflow_and_analytics(services):
    pid = _seed_finance_project(services)
    finance = services["finance_service"]
    reporting = services["reporting_service"]

    snapshot = finance.get_finance_snapshot(pid, as_of=date(2024, 1, 31), period="month")
    sources = {row.key: row for row in snapshot.by_source}
    policy = reporting.get_project_cost_source_breakdown(pid, as_of=date(2024, 1, 31))
    policy_map = {row.source_key: row for row in policy.rows}

    assert sources["DIRECT_COST"].actual == policy_map["DIRECT_COST"].actual
    assert sources["COMPUTED_LABOR"].actual == policy_map["COMPUTED_LABOR"].actual
    assert sources["LABOR_ADJUSTMENT"].actual == policy_map["LABOR_ADJUSTMENT"].actual

    manual_rows = [
        row
        for row in snapshot.ledger
        if row.source_key == "LABOR_ADJUSTMENT" and row.stage in {"planned", "actual"}
    ]
    assert manual_rows
    assert all(row.included_in_policy is False for row in manual_rows)

    assert snapshot.cashflow
    assert all(row.forecast == max(row.planned, row.committed) for row in snapshot.cashflow)
    assert all(row.exposure == max(row.committed, row.actual) for row in snapshot.cashflow)

    task_keys = {row.key for row in snapshot.by_task}
    assert "__project_level__" in task_keys
    assert any(key != "__project_level__" for key in task_keys)


def test_exporters_include_finance_sections_when_finance_service_is_provided(services, tmp_path):
    pid = _seed_finance_project(services)
    output_xlsx = tmp_path / "finance_report.xlsx"
    output_pdf = tmp_path / "finance_report.pdf"

    reporting_api.generate_excel_report(
        services["reporting_service"],
        pid,
        output_xlsx,
        finance_service=services["finance_service"],
        as_of=date(2024, 1, 31),
    )
    reporting_api.generate_pdf_report(
        services["reporting_service"],
        pid,
        output_pdf,
        finance_service=services["finance_service"],
        as_of=date(2024, 1, 31),
    )

    workbook = load_workbook(output_xlsx)
    assert "Finance" in workbook.sheetnames
    assert "Finance Ledger" in workbook.sheetnames
    assert workbook["Finance"]["A1"].value == "Finance Summary"
    assert workbook["Finance Ledger"]["A1"].value == "Date"

    assert output_pdf.exists()
    assert output_pdf.read_bytes().startswith(b"%PDF")


def test_finance_views_are_report_only_not_duplicated_in_cost_tab():
    root = Path(__file__).resolve().parents[1]
    cost_tab = (root / "ui" / "cost" / "tab.py").read_text(encoding="utf-8", errors="ignore")
    report_tab = (root / "ui" / "report" / "tab.py").read_text(encoding="utf-8", errors="ignore")
    report_actions = (root / "ui" / "report" / "actions.py").read_text(encoding="utf-8", errors="ignore")
    finance_dialog = (root / "ui" / "report" / "dialog_finance.py").read_text(
        encoding="utf-8",
        errors="ignore",
    )

    assert "finance_service: FinanceService | None = None" not in cost_tab
    assert "_build_finance_group" not in cost_tab
    assert "Show Finance View" in report_tab
    assert "def show_finance" in report_actions
    assert "finance_service=self._finance_service" in report_actions
    assert "self.main_splitter = QSplitter(Qt.Vertical)" in finance_dialog
    assert "self.main_splitter.addWidget(grp_cash)" in finance_dialog
    assert "self.main_splitter.addWidget(grp_analytics)" in finance_dialog
    assert "self.main_splitter.addWidget(grp_ledger)" in finance_dialog
