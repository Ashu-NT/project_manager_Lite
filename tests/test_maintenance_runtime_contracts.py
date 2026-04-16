from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from core.modules.maintenance_management.exporting import (
    MAINTENANCE_EXPORT_CONTRACTS,
    register_maintenance_management_export_definitions,
)
from core.modules.maintenance_management.importing import (
    MAINTENANCE_WORKBOOK_SHEETS,
    maintenance_import_schemas,
    maintenance_module_import_contracts,
    maintenance_reference_workbook_sheets,
    register_maintenance_management_import_definitions,
)
from core.modules.maintenance_management.reporting import (
    MAINTENANCE_REPORT_CONTRACTS,
    register_maintenance_management_report_definitions,
)
from core.modules.maintenance_management.services import MaintenanceRuntimeContractCatalogService
from core.platform.exporting import ExportDefinitionRegistry
from core.platform.importing import (
    ImportDefinitionRegistry,
    ImportPreview,
    ImportPreviewRow,
    ImportSummary,
)
from src.core.platform.report_runtime import ReportDefinitionRegistry


def _preview(rows):
    return ImportPreview(
        entity_type="maintenance",
        available_columns=[],
        mapped_columns={},
        rows=[
            ImportPreviewRow(
                line_no=rows[0].line_no if rows else 1,
                status="READY",
                action="CREATE",
                message="Ready",
                row=dict(rows[0].values) if rows else {},
            )
        ],
        created_count=len(rows),
    )


def _execute(rows):
    return ImportSummary(entity_type="maintenance", created_count=len(rows))


def test_maintenance_workbook_contracts_split_module_owned_and_shared_reference_sheets() -> None:
    module_owned = {sheet.sheet_name for sheet in maintenance_module_import_contracts()}
    shared_refs = {sheet.sheet_name for sheet in maintenance_reference_workbook_sheets()}
    ownership = {sheet.sheet_name: sheet.owner_module_code for sheet in MAINTENANCE_WORKBOOK_SHEETS}

    assert module_owned == {
        "Locations",
        "Systems",
        "Assets",
        "AssetComponents",
        "Sensors",
        "TaskTemplates",
        "TaskStepTemplates",
        "PreventivePlans",
        "PreventivePlanTasks",
        "DocumentLinks",
    }
    assert shared_refs == {
        "Parties",
        "StockItems",
        "Storerooms",
        "StockBalances",
        "Documents",
        "DocumentStructures",
    }
    assert ownership["Parties"] == "platform"
    assert ownership["Documents"] == "platform"
    assert ownership["StockItems"] == "inventory_procurement"
    assert ownership["DocumentLinks"] == "maintenance_management"


def test_maintenance_import_contracts_expose_expected_workbook_schemas() -> None:
    schemas = maintenance_import_schemas()

    assert set(schemas) == {
        "maintenance_locations",
        "maintenance_systems",
        "maintenance_assets",
        "maintenance_asset_components",
        "maintenance_sensors",
        "maintenance_task_templates",
        "maintenance_task_step_templates",
        "maintenance_preventive_plans",
        "maintenance_preventive_plan_tasks",
        "maintenance_document_links",
    }
    assert schemas["maintenance_assets"][0].key == "asset_code"
    assert schemas["maintenance_preventive_plan_tasks"][0].key == "plan_code"


def test_maintenance_import_definition_registration_uses_module_owned_contracts() -> None:
    registry = ImportDefinitionRegistry()
    handlers = {
        contract.operation_key: _preview
        for contract in maintenance_module_import_contracts()
        if contract.operation_key
    }
    execute_handlers = {
        contract.operation_key: _execute
        for contract in maintenance_module_import_contracts()
        if contract.operation_key
    }

    register_maintenance_management_import_definitions(
        registry,
        preview_handlers=handlers,
        execution_handlers=execute_handlers,
    )

    assert set(registry.list_operation_keys()) == set(handlers)
    definition = registry.get("maintenance_assets")
    assert definition.module_code == "maintenance_management"
    assert definition.permission_code == "import.manage"


def test_maintenance_export_contract_registration_exposes_known_operation_keys() -> None:
    registry = ExportDefinitionRegistry()
    handlers = {
        contract.operation_key: (lambda request, key=contract.operation_key: Path(f"{key}.csv"))
        for contract in MAINTENANCE_EXPORT_CONTRACTS
    }

    register_maintenance_management_export_definitions(registry, export_handlers=handlers)

    assert set(registry.list_operation_keys()) == {
        contract.operation_key for contract in MAINTENANCE_EXPORT_CONTRACTS
    }
    definition = registry.get("maintenance_asset_register_excel")
    assert definition.module_code == "maintenance_management"
    assert definition.permission_code == "report.export"


def test_maintenance_report_contract_registration_exposes_known_report_keys() -> None:
    registry = ReportDefinitionRegistry()
    handlers = {
        contract.report_key: (lambda request, key=contract.report_key: {"report_key": key, "request": request})
        for contract in MAINTENANCE_REPORT_CONTRACTS
    }

    register_maintenance_management_report_definitions(registry, render_handlers=handlers)

    assert set(registry.list_report_keys()) == {
        contract.report_key for contract in MAINTENANCE_REPORT_CONTRACTS
    }
    definition = registry.get("maintenance_downtime_pdf")
    assert definition.module_code == "maintenance_management"
    assert definition.supported_formats == ("pdf",)


def test_maintenance_runtime_contract_catalog_service_exports_rollout_workbook_and_matrix(tmp_path) -> None:
    service = MaintenanceRuntimeContractCatalogService()

    workbook_artifact = service.export_rollout_workbook_template(tmp_path / "maintenance-rollout-template.xlsx")
    matrix_artifact = service.export_runtime_contract_matrix(tmp_path / "maintenance-contract-matrix.csv")

    workbook = load_workbook(workbook_artifact.file_path)
    matrix_text = matrix_artifact.file_path.read_text(encoding="utf-8")

    assert workbook_artifact.metadata["artifact_kind"] == "maintenance_rollout_workbook_template"
    assert "Workbook Index" in workbook.sheetnames
    assert "Locations" in workbook.sheetnames
    assert workbook["Locations"]["A1"].value == "Maintenance Locations"
    assert workbook["Locations"]["A7"].value == "location_code"
    assert matrix_artifact.metadata["artifact_kind"] == "maintenance_runtime_contract_matrix"
    assert "maintenance_locations" in matrix_text
    assert "maintenance_asset_register_excel" in matrix_text
    assert "maintenance_backlog_pdf" in matrix_text


def test_maintenance_runtime_contract_catalog_service_builds_report_document() -> None:
    service = MaintenanceRuntimeContractCatalogService()

    report = service.build_contract_catalog_report()

    assert report.title == "Maintenance Runtime Contract Catalog"
    assert [section.title for section in report.sections] == [
        "Summary",
        "Workbook",
        "Exports",
        "Reports",
    ]
    summary_metric_block = report.sections[0].blocks[0]
    workbook_table = report.sections[1].blocks[0]
    assert summary_metric_block.rows[0].value == len(MAINTENANCE_WORKBOOK_SHEETS)
    assert workbook_table.rows[0][0] == "Locations"
