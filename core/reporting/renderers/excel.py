from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from core.reporting.contexts import ExcelReportContext


class ExcelReportRenderer:
    def render(self, ctx: ExcelReportContext, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

        header_font = Font(bold=True)
        title_font = Font(bold=True, size=14)
        center = Alignment(horizontal="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        header_fill = PatternFill("solid", fgColor="DDDDDD")

        # ---------------- Overview ----------------
        ws = wb.active
        ws.title = "Overview"

        ws["A1"] = f"Project KPIs - {ctx.kpi.name}"
        ws["A1"].font = title_font

        row = 3

        def kv(key, value):
            nonlocal row
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = header_font
            ws[f"A{row}"].border = thin_border
            ws[f"B{row}"].border = thin_border
            row += 1

        kv("Project ID", ctx.kpi.project_id)
        kv("Project name", ctx.kpi.name)
        kv("Start date", ctx.kpi.start_date)
        kv("End date", ctx.kpi.end_date)
        kv("Duration (working days)", ctx.kpi.duration_working_days)

        row += 1
        kv("Tasks - total", ctx.kpi.tasks_total)
        kv("Tasks - completed", ctx.kpi.tasks_completed)
        kv("Tasks - in progress", ctx.kpi.tasks_in_progress)
        kv("Tasks - not started", ctx.kpi.tasks_not_started)
        kv("Critical tasks", ctx.kpi.critical_tasks)
        kv("Late tasks", ctx.kpi.late_tasks)

        row += 1
        kv("Planned cost", ctx.kpi.total_planned_cost)
        kv("Actual cost", ctx.kpi.total_actual_cost)
        kv("Cost variance", ctx.kpi.cost_variance)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

        # ---------------- Tasks ----------------
        ws_tasks = wb.create_sheet("Tasks")
        headers = ["Task ID", "Name", "Start", "End", "Duration (days)", "Critical", "% complete", "Status"]
        for col_index, h in enumerate(headers, start=1):
            cell = ws_tasks.cell(row=1, column=col_index, value=h)
            cell.font = header_font
            cell.alignment = center
            cell.fill = header_fill
            cell.border = thin_border

        for row_index, b in enumerate(ctx.gantt, start=2):
            ws_tasks.cell(row=row_index, column=1, value=b.task_id).border = thin_border
            ws_tasks.cell(row=row_index, column=2, value=b.name).border = thin_border
            ws_tasks.cell(row=row_index, column=3, value=b.start.isoformat() if b.start else "").border = thin_border
            ws_tasks.cell(row=row_index, column=4, value=b.end.isoformat() if b.end else "").border = thin_border
            dur = (b.end - b.start).days + 1 if (b.start and b.end) else None
            ws_tasks.cell(row=row_index, column=5, value=dur).border = thin_border
            ws_tasks.cell(row=row_index, column=6, value="Yes" if b.is_critical else "No").border = thin_border
            ws_tasks.cell(row=row_index, column=7, value=b.percent_complete).border = thin_border
            ws_tasks.cell(row=row_index, column=8, value=getattr(b.status, "value", str(b.status))).border = thin_border

        ws_tasks.column_dimensions["A"].width = 36
        ws_tasks.column_dimensions["B"].width = 30
        for col_letter in ("C", "D", "E", "F", "G", "H"):
            ws_tasks.column_dimensions[col_letter].width = 15

        # ---------------- Resources ----------------
        ws_res = wb.create_sheet("Resources")
        headers = ["Resource ID", "Name", "Total allocation (%)", "Tasks count"]
        for c, h in enumerate(headers, start=1):
            cell = ws_res.cell(1, c, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for r, res in enumerate(ctx.resources, start=2):
            values = [
                res.resource_id,
                res.resource_name,
                res.total_allocation_percent,
                res.tasks_count,
            ]
            for c, v in enumerate(values, 1):
                cell = ws_res.cell(r, c, v)
                cell.border = thin_border

        ws_res.column_dimensions["A"].width = 34
        ws_res.column_dimensions["B"].width = 28
        ws_res.column_dimensions["C"].width = 20
        ws_res.column_dimensions["D"].width = 15

        # ---------------- EVM ----------------
        if ctx.evm or ctx.evm_series:
            ws_evm = wb.create_sheet("EVM")
            ws_evm["A1"] = "Earned Value Management"
            ws_evm["A1"].font = title_font

            if ctx.evm:
                ws_evm["A2"], ws_evm["B2"] = ("Metric", "Value")
                ws_evm["A2"].font = header_font
                ws_evm["B2"].font = header_font
                ws_evm["A2"].fill = header_fill
                ws_evm["B2"].fill = header_fill
                ws_evm["A2"].border = thin_border
                ws_evm["B2"].border = thin_border

                rows = [
                    ("As of", ctx.as_of.isoformat()),
                    ("BAC", ctx.evm.BAC),
                    ("PV", ctx.evm.PV),
                    ("EV", ctx.evm.EV),
                    ("AC", ctx.evm.AC),
                    ("SPI", ctx.evm.SPI),
                    ("CPI", ctx.evm.CPI),
                    ("EAC", ctx.evm.EAC),
                    ("ETC", ctx.evm.ETC),
                    ("VAC", ctx.evm.VAC),
                ]

                for i, (k, v) in enumerate(rows, start=3):
                    ws_evm[f"A{i}"] = k
                    if isinstance(v, (int, float)):
                        ws_evm[f"B{i}"] = float(v)
                    elif v is None:
                        ws_evm[f"B{i}"] = ""
                    else:
                        ws_evm[f"B{i}"] = str(v)
                    ws_evm[f"A{i}"].border = thin_border
                    ws_evm[f"B{i}"].border = thin_border

                ws_evm.column_dimensions["A"].width = 22
                ws_evm.column_dimensions["B"].width = 18

            if ctx.evm_series:
                ws_evm["D2"] = "Period End"
                ws_evm["E2"] = "PV"
                ws_evm["F2"] = "EV"
                ws_evm["G2"] = "AC"
                for col in ("D", "E", "F", "G"):
                    ws_evm[f"{col}2"].font = header_font
                    ws_evm[f"{col}2"].fill = header_fill
                    ws_evm[f"{col}2"].border = thin_border

                for idx, point in enumerate(ctx.evm_series, start=3):
                    ws_evm[f"D{idx}"] = point.period_end.isoformat()
                    ws_evm[f"E{idx}"] = float(point.PV or 0.0)
                    ws_evm[f"F{idx}"] = float(point.EV or 0.0)
                    ws_evm[f"G{idx}"] = float(point.AC or 0.0)
                    for col in ("D", "E", "F", "G"):
                        ws_evm[f"{col}{idx}"].border = thin_border

                ws_evm.column_dimensions["D"].width = 14
                ws_evm.column_dimensions["E"].width = 14
                ws_evm.column_dimensions["F"].width = 14
                ws_evm.column_dimensions["G"].width = 14

        # ---------------- Baseline Variance ----------------
        if ctx.baseline_variance:
            ws_v = wb.create_sheet("Variance")
            headers = ["Task", "Baseline Start", "Baseline Finish", "Current Start", "Current Finish", "SV days", "FV days", "Critical"]
            for i, h in enumerate(headers, start=1):
                cell = ws_v.cell(row=1, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

            for r_i, row in enumerate(ctx.baseline_variance, start=2):
                ws_v.cell(r_i, 1, row.task_name).border = thin_border
                ws_v.cell(r_i, 2, "" if not row.baseline_start else row.baseline_start.isoformat()).border = thin_border
                ws_v.cell(r_i, 3, "" if not row.baseline_finish else row.baseline_finish.isoformat()).border = thin_border
                ws_v.cell(r_i, 4, "" if not row.current_start else row.current_start.isoformat()).border = thin_border
                ws_v.cell(r_i, 5, "" if not row.current_finish else row.current_finish.isoformat()).border = thin_border
                ws_v.cell(r_i, 6, "" if row.start_variance_days is None else row.start_variance_days).border = thin_border
                ws_v.cell(r_i, 7, "" if row.finish_variance_days is None else row.finish_variance_days).border = thin_border
                ws_v.cell(r_i, 8, "Yes" if row.is_critical else "No").border = thin_border

            ws_v.column_dimensions["A"].width = 40
            for col in ("B", "C", "D", "E", "F", "G", "H"):
                ws_v.column_dimensions[col].width = 15

        # ---------------- Cost Breakdown ----------------
        if ctx.cost_breakdown:
            ws_c = wb.create_sheet("Cost Breakdown")
            headers = ["Type", "Currency", "Planned", "Actual", "Variance"]
            for i, h in enumerate(headers, start=1):
                cell = ws_c.cell(row=1, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

            for r_i, row in enumerate(ctx.cost_breakdown, start=2):
                planned = float(row.planned or 0.0)
                actual = float(row.actual or 0.0)
                ws_c.cell(r_i, 1, str(row.cost_type)).border = thin_border
                ws_c.cell(r_i, 2, str(row.currency)).border = thin_border
                ws_c.cell(r_i, 3, planned).border = thin_border
                ws_c.cell(r_i, 4, actual).border = thin_border
                ws_c.cell(r_i, 5, actual - planned).border = thin_border

            ws_c.column_dimensions["A"].width = 18
            ws_c.column_dimensions["B"].width = 10
            ws_c.column_dimensions["C"].width = 14
            ws_c.column_dimensions["D"].width = 14
            ws_c.column_dimensions["E"].width = 14

        wb.save(output_path)
        return output_path
