# reporting/exporters.py
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
from typing import List

import matplotlib.pyplot as plt
from matplotlib.dates import date2num
import matplotlib.dates as mdates
from matplotlib import ticker

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as PdfImage,
)
from reportlab.lib.styles import getSampleStyleSheet

from core.services.reporting import ReportingService
from core.services.reporting import GanttTaskBar, ProjectKPI, ResourceLoadRow

def _ensure_path(path: str | Path) -> Path:
    p = Path(path)
    if not p.parent.exists():
        p.parent.mkdir(parents=True, exist_ok=True)
    return p

def generate_gantt_png(
    reporting_service: ReportingService,
    project_id: str,
    output_path: str | Path,
) -> Path:
    """
    Generate a Gantt chart as a PNG image for a project.
    Uses matplotlib, styled for readability and a clean date axis.
    """
    output_path = _ensure_path(output_path)

    bars: List[GanttTaskBar] = reporting_service.get_gantt_data(project_id)
    bars = [b for b in bars if b.start and b.end]
    if not bars:
        raise ValueError("No tasks with dates available for Gantt chart.")

    # Sort tasks by start date
    bars.sort(key=lambda b: (b.start, b.end or b.start))

    names = [b.name for b in bars]
    start_nums = [date2num(b.start) for b in bars]
    #durations = [(b.end - b.start).days + 1 for b in bars]
    durations = [(b.end - b.start).days + 0 for b in bars]
    critical_flags = [b.is_critical for b in bars]
    percents = [b.percent_complete or 0.0 for b in bars]

    fig, ax = plt.subplots(figsize=(12, 6))
    y_positions = range(len(bars))

    # Draw bars
    for i, (start, dur, is_crit, pct) in enumerate(zip(start_nums, durations, critical_flags, percents)):
        base_color = "#d0d0ff" if not is_crit else "#ffcccc"
        ax.barh(
            y_positions[i],
            dur,
            left=start,
            height=0.4,
            color=base_color,
            edgecolor="black",
            linewidth=0.6,
        )

        if pct > 0:
            completed_dur = dur * pct / 100.0
            comp_color = "#8080ff" if not is_crit else "#ff6666"
            ax.barh(
                y_positions[i],
                completed_dur,
                left=start,
                height=0.4,
                color=comp_color,
                edgecolor=None,
            )

    # Y axis
    ax.set_yticks(list(y_positions))
    ax.set_yticklabels(names, fontsize=9)
    ax.invert_yaxis()

    # X axis as dates with a single, clean set of labels
    locator = mdates.AutoDateLocator(minticks=4, maxticks=10)
    formatter = mdates.ConciseDateFormatter(locator)   # e.g. 2025-12-09 shown nicely

    ax.xaxis.set_major_locator(locator)
    ax.xaxis.set_major_formatter(formatter)

    # IMPORTANT: disable minor tick labels to avoid duplicates
    ax.xaxis.set_minor_locator(ticker.NullLocator())
    ax.xaxis.set_minor_formatter(ticker.NullFormatter())

    fig.autofmt_xdate(rotation=30)

    # Optional today line (keep or remove to taste)
   
    today = date.today()
    
    # Determine project range
    start_dates = [ b.start for b in bars if b.start]
    end_dates = [b.end for b in bars if b.end]
    
    if start_dates and end_dates:
        project_start = min(start_dates)
        project_end = max(end_dates)
    else:
        project_start = project_end = None
    
    # only draw "Today" Line if today lies withing the project timeline (with 7 - day pad)
    if project_start and project_end:
        buffer = 7 # days buffer on either side
        if project_start - timedelta(days=buffer) <= today <= project_end + timedelta(days=buffer):
    
            today_num = date2num(today)
            ax.axvline(today_num, color="red", linestyle="--", linewidth=1, alpha=0.7)
            ax.text(
                today_num,
                len(bars) + 0.1,
                "Today",
                color="red",
                fontsize=8,
                rotation=90,
                ha="center",
                va="bottom",
            )

    ax.set_title("Project Gantt Chart", fontsize=14, pad=10)
    ax.set_xlabel("Date")
    ax.grid(True, axis="x", linestyle=":", linewidth=0.5)

    plt.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)

    return output_path

def generate_evm_png(
    reporting_service: ReportingService,
    project_id: str,
    output_path: str | Path,
    baseline_id: str | None = None,
    as_of: date | None = None,
) -> Path:
    """
    Generates an EVM S-curve (PV/EV/AC over time) PNG.
    Uses reporting_service.get_evm_series() if available.
    """
    output_path = _ensure_path(output_path)
    as_of = as_of or date.today()

    if not hasattr(reporting_service, "get_evm_series"):
        return output_path  # gracefully do nothing

    series = reporting_service.get_evm_series(project_id, baseline_id=baseline_id, as_of=as_of)
    if not series:
        return output_path

    xs = [p.period_end for p in series]
    pv = [float(p.PV or 0.0) for p in series]
    ev = [float(p.EV or 0.0) for p in series]
    ac = [float(p.AC or 0.0) for p in series]

    fig, ax = plt.subplots(figsize=(9, 3))
    ax.plot(xs, pv, label="PV")
    ax.plot(xs, ev, label="EV")
    ax.plot(xs, ac, label="AC")
    ax.set_title("EVM S-Curve (Cumulative)")
    ax.set_xlabel("Period")
    ax.set_ylabel("Cost")
    ax.grid(True, axis="y", linestyle=":", linewidth=0.6)
    ax.legend()

    plt.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)
    return output_path

def generate_excel_report(
    reporting_service: ReportingService,
    project_id: str,
    output_path: str | Path,
) -> Path:
    """
    Generate an Excel report (.xlsx) for a project:
    - Overview: KPIs
    - Tasks: Gantt data
    - Resources: resource load summary
    """
    output_path = _ensure_path(output_path)

    kpi: ProjectKPI = reporting_service.get_project_kpis(project_id)
    as_of = date.today()
    baseline_id = None
    gantt_data: List[GanttTaskBar] = reporting_service.get_gantt_data(project_id)
    res_load: List[ResourceLoadRow] = reporting_service.get_resource_load_summary(project_id)

    wb = Workbook()

    # Styles
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill("solid", fgColor="DDDDDD")

    # --- Overview sheet --- #
    ws = wb.active
    ws.title = "Overview"

    ws["A1"] = f"Project KPIs - {kpi.name}"
    ws["A1"].font = title_font

    row = 3
    def kv(key, value):
        nonlocal row
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        ws[f"A{row}"].font = header_font
        row += 1

    kv("Project ID", kpi.project_id)
    kv("Project name", kpi.name)
    kv("Start date", kpi.start_date.isoformat() if kpi.start_date else "")
    kv("End date", kpi.end_date.isoformat() if kpi.end_date else "")
    kv("Duration (working days)", kpi.duration_working_days)

    row += 1
    kv("Tasks - total", kpi.tasks_total)
    kv("Tasks - completed", kpi.tasks_completed)
    kv("Tasks - in progress", kpi.tasks_in_progress)
    kv("Tasks - not started", kpi.tasks_not_started)
    kv("Critical tasks", kpi.critical_tasks)
    kv("Late tasks", kpi.late_tasks)

    row += 1
    kv("Planned cost", kpi.total_planned_cost)
    kv("Actual cost", kpi.total_actual_cost)
    kv("Cost variance", kpi.cost_variance)

    # Apply borders & widths roughly
    for r in range(3, row):
        for col in ("A", "B"):
            cell = ws[f"{col}{r}"]
            cell.border = thin_border
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    # --- Tasks sheet --- #
    ws_tasks = wb.create_sheet("Tasks")
    headers = ["Task ID", "Name", "Start", "End", "Duration (days)", "Critical", "% complete", "Status"]
    for col_index, h in enumerate(headers, start=1):
        cell = ws_tasks.cell(row=1, column=col_index, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    for row_index, b in enumerate(gantt_data, start=2):
        ws_tasks.cell(row=row_index, column=1, value=b.task_id).border = thin_border
        ws_tasks.cell(row=row_index, column=2, value=b.name).border = thin_border
        ws_tasks.cell(row=row_index, column=3, value=b.start.isoformat() if b.start else "").border = thin_border
        ws_tasks.cell(row=row_index, column=4, value=b.end.isoformat() if b.end else "").border = thin_border
        dur = (b.end - b.start).days + 1 if (b.start and b.end) else None
        ws_tasks.cell(row=row_index, column=5, value=dur).border = thin_border
        ws_tasks.cell(row=row_index, column=6, value="Yes" if b.is_critical else "No").border = thin_border
        ws_tasks.cell(row=row_index, column=7, value=b.percent_complete).border = thin_border
        ws_tasks.cell(row=row_index, column=8, value=getattr(b.status, "value", str(b.status))).border = thin_border

    ws_tasks.column_dimensions["A"].width = 36
    ws_tasks.column_dimensions["B"].width = 30
    for col_letter in ("C", "D", "E", "F", "G", "H"):
        ws_tasks.column_dimensions[col_letter].width = 15

    # --- Resources sheet --- #
    ws_res = wb.create_sheet("Resources")
    headers = ["Resource ID", "Name", "Total allocation (%)", "Tasks count"]
    for col_index, h in enumerate(headers, start=1):
        cell = ws_res.cell(row=1, column=col_index, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = thin_border

    for row_index, r in enumerate(res_load, start=2):
        ws_res.cell(row=row_index, column=1, value=r.resource_id).border = thin_border
        ws_res.cell(row=row_index, column=2, value=r.resource_name).border = thin_border
        ws_res.cell(row=row_index, column=3, value=r.total_allocation_percent).border = thin_border
        ws_res.cell(row=row_index, column=4, value=r.tasks_count).border = thin_border

    ws_res.column_dimensions["A"].width = 34
    ws_res.column_dimensions["B"].width = 28
    ws_res.column_dimensions["C"].width = 20
    ws_res.column_dimensions["D"].width = 15
    
    # --- EVM sheet ---
    if hasattr(reporting_service, "get_earned_value"):
        ws_evm = wb.create_sheet("EVM")
        ws_evm["A1"] = "Earned Value Management (EVM)"
        ws_evm["A1"].font = title_font

        evm = reporting_service.get_earned_value(
            project_id,
            baseline_id=baseline_id,
            as_of=as_of
        )

        rows = [
            ("As of", as_of.isoformat()),
            ("BAC", float(getattr(evm, "BAC", 0.0) or 0.0)),
            ("PV",  float(getattr(evm, "PV", 0.0) or 0.0)),
            ("EV",  float(getattr(evm, "EV", 0.0) or 0.0)),
            ("AC",  float(getattr(evm, "AC", 0.0) or 0.0)),
            ("ETC",  float(getattr(evm, "ETC", 0.0) or 0.0)),
            ("TCPI_to_BAC",  float(getattr(evm, "TCPI_to_BAC", 0.0) or 0.0)),
            ("TCPI_to_EAC",  float(getattr(evm, "TCPI_to_EAC", 0.0) or 0.0)),
            ("SPI", float(getattr(evm, "SPI", 0.0) or 0.0)),
            ("CPI", float(getattr(evm, "CPI", 0.0) or 0.0)),
            ("EAC", float(getattr(evm, "EAC", 0.0) or 0.0)),
            ("VAC", float(getattr(evm, "VAC", 0.0) or 0.0)),
        ]

        r0 = 3
        ws_evm["A2"] = "Metric"; ws_evm["B2"] = "Value"
        ws_evm["A2"].font = header_font; ws_evm["B2"].font = header_font
        ws_evm["A2"].fill = header_fill; ws_evm["B2"].fill = header_fill
        ws_evm["A2"].border = thin_border; ws_evm["B2"].border = thin_border

        for i, (k, v) in enumerate(rows):
            rr = r0 + i
            ws_evm[f"A{rr}"] = k
            ws_evm[f"B{rr}"] = v
            ws_evm[f"A{rr}"].border = thin_border
            ws_evm[f"B{rr}"].border = thin_border

        ws_evm.column_dimensions["A"].width = 22
        ws_evm.column_dimensions["B"].width = 18

    if hasattr(reporting_service, "get_evm_series"):
        series = reporting_service.get_evm_series(project_id, baseline_id=baseline_id, as_of=as_of)
        if series:
            ws_evm["D2"] = "Period End"
            ws_evm["E2"] = "PV"
            ws_evm["F2"] = "EV"
            ws_evm["G2"] = "AC"
            for c in ("D","E","F","G"):
                ws_evm[f"{c}2"].font = header_font
                ws_evm[f"{c}2"].fill = header_fill
                ws_evm[f"{c}2"].border = thin_border

            for idx, p in enumerate(series, start=3):
                ws_evm[f"D{idx}"] = p.period_end.isoformat()
                ws_evm[f"E{idx}"] = float(p.PV or 0.0)
                ws_evm[f"F{idx}"] = float(p.EV or 0.0)
                ws_evm[f"G{idx}"] = float(p.AC or 0.0)
                for c in ("D","E","F","G"):
                    ws_evm[f"{c}{idx}"].border = thin_border

            ws_evm.column_dimensions["D"].width = 14
            for c in ("E","F","G"):
                ws_evm.column_dimensions[c].width = 14
      
    if hasattr(reporting_service, "get_baseline_schedule_variance"):
        rows = reporting_service.get_baseline_schedule_variance(project_id, baseline_id=baseline_id)
        if rows:
            ws_v = wb.create_sheet("Variance")
            headers = ["Task", "Baseline Start", "Baseline Finish", "Current Start", "Current Finish", "SV days", "FV days", "Critical"]
            for i, h in enumerate(headers, start=1):
                cell = ws_v.cell(row=1, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

            for r_i, r in enumerate(rows, start=2):
                ws_v.cell(r_i, 1, r.task_name).border = thin_border
                ws_v.cell(r_i, 2, "" if not r.baseline_start else r.baseline_start.isoformat()).border = thin_border
                ws_v.cell(r_i, 3, "" if not r.baseline_finish else r.baseline_finish.isoformat()).border = thin_border
                ws_v.cell(r_i, 4, "" if not r.current_start else r.current_start.isoformat()).border = thin_border
                ws_v.cell(r_i, 5, "" if not r.current_finish else r.current_finish.isoformat()).border = thin_border
                ws_v.cell(r_i, 6, "" if r.start_variance_days is None else r.start_variance_days).border = thin_border
                ws_v.cell(r_i, 7, "" if r.finish_variance_days is None else r.finish_variance_days).border = thin_border
                ws_v.cell(r_i, 8, "Yes" if r.is_critical else "No").border = thin_border

            ws_v.column_dimensions["A"].width = 40
            for c in ("B","C","D","E","F","G","H"):
                ws_v.column_dimensions[c].width = 15
    
    if hasattr(reporting_service, "get_cost_breakdown"):
        rows = reporting_service.get_cost_breakdown(project_id, as_of=as_of)
        if rows:
            ws_c = wb.create_sheet("Cost Breakdown")
            headers = ["Type", "Currency", "Planned", "Actual", "Variance"]
            for i, h in enumerate(headers, start=1):
                cell = ws_c.cell(row=1, column=i, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

            for r_i, r in enumerate(rows, start=2):
                planned = float(r.planned or 0.0)
                actual = float(r.actual or 0.0)
                ws_c.cell(r_i, 1, str(r.cost_type)).border = thin_border
                ws_c.cell(r_i, 2, str(r.currency)).border = thin_border
                ws_c.cell(r_i, 3, planned).border = thin_border
                ws_c.cell(r_i, 4, actual).border = thin_border
                ws_c.cell(r_i, 5, actual - planned).border = thin_border

            ws_c.column_dimensions["A"].width = 18
            ws_c.column_dimensions["B"].width = 10
            for c in ("C","D","E"):
                ws_c.column_dimensions[c].width = 14
    
    wb.save(output_path)
    return output_path

def generate_pdf_report(
    reporting_service: ReportingService,
    project_id: str,
    output_path: str | Path,
    temp_dir: str | Path = "tmp_reports",
) -> Path:
    """
    Generate a PDF report:
    - Title & basic info
    - KPIs table
    - Gantt chart image
    - Resource load table
    """
    output_path = _ensure_path(output_path)
    temp_dir = Path(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)

    kpi: ProjectKPI = reporting_service.get_project_kpis(project_id)
    
    as_of = date.today()
    baseline_id = None  
    
    gantt_png = temp_dir / f"gantt_{project_id}.png"
    gantt_png = generate_gantt_png(reporting_service, project_id, gantt_png)

    res_load: List[ResourceLoadRow] = reporting_service.get_resource_load_summary(project_id)

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=landscape(A4),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title
    title = f"Project Report - {kpi.name}"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 12))

    # Basic info
    info_lines = [
        f"Project ID: {kpi.project_id}",
        f"Start date: {kpi.start_date or '-'}",
        f"End date: {kpi.end_date or '-'}",
        f"Duration (working days): {kpi.duration_working_days}",
        "",
        f"Tasks: total={kpi.tasks_total}, completed={kpi.tasks_completed}, "
        f"in progress={kpi.tasks_in_progress}, not started={kpi.tasks_not_started}",
        f"Critical tasks: {kpi.critical_tasks}, Late tasks: {kpi.late_tasks}",
        "",
        f"Cost: planned={kpi.total_planned_cost:.2f}, "
        f"actual={kpi.total_actual_cost:.2f}, "
        f"variance={kpi.cost_variance:.2f}",
    ]
    for line in info_lines:
        story.append(Paragraph(line, styles["Normal"]))
    story.append(Spacer(1, 16))

    # KPIs table for a quick visual
    kpi_data = [
        ["Metric", "Value"],
        ["Start date", str(kpi.start_date or "-")],
        ["End date", str(kpi.end_date or "-")],
        ["Duration (working days)", kpi.duration_working_days],
        ["Total tasks", kpi.tasks_total],
        ["Completed", kpi.tasks_completed],
        ["In progress", kpi.tasks_in_progress],
        ["Not started", kpi.tasks_not_started],
        ["Critical tasks", kpi.critical_tasks],
        ["Late tasks", kpi.late_tasks],
        ["Planned cost", f"{kpi.total_planned_cost:.2f}"],
        ["Actual cost", f"{kpi.total_actual_cost:.2f}"],
        ["Cost variance", f"{kpi.cost_variance:.2f}"],
    ]
    tbl = Table(kpi_data, hAlign="LEFT", colWidths=[140, 150])
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )
    story.append(tbl)
    story.append(Spacer(1, 20))

# --- EVM Summary (professional section) ---
    if hasattr(reporting_service, "get_earned_value"):
        evm = reporting_service.get_earned_value(
            project_id,
            baseline_id=baseline_id,
            as_of=as_of
        )

        story.append(Paragraph("Earned Value Management (EVM)", styles["Heading2"]))
        story.append(Spacer(1, 8))
        
        evm_rows = [
            ["As of", as_of.isoformat()],
            ["BAC", f"{float(getattr(evm, 'BAC', 0.0) or 0.0):.2f}"],
            ["PV",  f"{float(getattr(evm, 'PV',  0.0)or 0.0):.2f}"],
            ["EV",  f"{float(getattr(evm, 'EV',  0.0)or 0.0):.2f}"],
            ["AC",  f"{float(getattr(evm, 'AC',  0.0)or 0.0):.2f}"],
            ["SV (EV − PV)", f"{float(getattr(evm, 'SV', 0.0)or 0.0):.2f}"],
            ["CV (EV − AC)", f"{float(getattr(evm, 'CV', 0.0)or 0.0):.2f}"],
            ["SPI", f"{float(getattr(evm, 'SPI', 0.0)or 0.0):.3f}"],
            ["CPI", f"{float(getattr(evm, 'CPI', 0.0)or 0.0):.3f}"],
            ["EAC", f"{float(getattr(evm, 'EAC', 0.0)or 0.0):.2f}"],
            ["VAC (BAC − EAC)", f"{float(getattr(evm, 'VAC', 0.0)or 0.0):.2f}"],
        ]

        evm_tbl = Table([["Metric", "Value"]] + evm_rows, hAlign="LEFT", colWidths=[180, 160])
        evm_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ]))
        story.append(evm_tbl)
        story.append(Spacer(1, 16))

    # --- Baseline schedule variance ---
    if hasattr(reporting_service, "get_baseline_schedule_variance"):
        rows = reporting_service.get_baseline_schedule_variance(project_id, baseline_id=baseline_id)
        if rows:
            story.append(Paragraph("Baseline vs Current Schedule Variance (Top 15)", styles["Heading2"]))
            story.append(Spacer(1, 8))

            data = [["Task", "Baseline Finish", "Current Finish", "Finish Var (days)", "Critical"]]
            for r in rows[:15]:
                data.append([
                    r.task_name,
                    str(r.baseline_finish or "-"),
                    str(r.current_finish or "-"),
                    "" if r.finish_variance_days is None else str(r.finish_variance_days),
                    "Yes" if getattr(r, "is_critical", False) else "No",
                ])

            t = Table(data, hAlign="LEFT", colWidths=[320, 110, 110, 110, 70])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ]))
            story.append(t)
            story.append(Spacer(1, 16))

    # --- Cost breakdown ---
    if hasattr(reporting_service, "get_cost_breakdown"):
        cb = reporting_service.get_cost_breakdown(project_id, as_of=as_of)
        if cb:
            story.append(Paragraph("Cost Breakdown (Planned vs Actual)", styles["Heading2"]))
            story.append(Spacer(1, 8))

            data = [["Type", "Currency", "Planned", "Actual", "Variance"]]
            for r in cb:
                var = float(r.actual or 0.0) - float(r.planned or 0.0)
                data.append([
                    str(r.cost_type),
                    str(r.currency),
                    f"{float(r.planned or 0.0):.2f}",
                    f"{float(r.actual or 0.0):.2f}",
                    f"{var:.2f}",
                ])

            t = Table(data, hAlign="LEFT", colWidths=[140, 80, 120, 120, 120])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ]))
            story.append(t)
            story.append(Spacer(1, 16))
        
    # Gantt image
    story.append(Paragraph("Gantt Chart", styles["Heading2"]))
    story.append(Spacer(1, 8))

    img = PdfImage(str(gantt_png))
    img._restrictSize(720, 280)  # fit on page
    story.append(img)
    story.append(Spacer(1, 20))

    # Resource load table
    if res_load:
        story.append(Paragraph("Resource Load Summary", styles["Heading2"]))
        story.append(Spacer(1, 8))

        res_data = [["Resource", "Total allocation (%)", "Tasks count"]]
        for r in res_load:
            res_data.append(
                [f"{r.resource_name} ({r.resource_id})", f"{r.total_allocation_percent:.1f}", r.tasks_count]
            )

        res_tbl = Table(res_data, hAlign="LEFT", colWidths=[260, 140, 100])
        res_tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
                ]
            )
        )
        story.append(res_tbl)

    doc.build(story)
    return output_path



