from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.modules.maintenance_management.exporting import MAINTENANCE_EXPORT_CONTRACTS
from core.modules.maintenance_management.importing import MAINTENANCE_WORKBOOK_SHEETS
from core.modules.maintenance_management.reporting import MAINTENANCE_REPORT_CONTRACTS
from src.core.platform.exporting import ensure_output_path, finalize_artifact
from src.core.platform.report_runtime import MetricBlock, MetricRow, ReportDocument, ReportSection, TableBlock, TextBlock


class MaintenanceRuntimeContractCatalogService:
    def __init__(self) -> None:
        self._header_font = Font(bold=True)
        self._title_font = Font(bold=True, size=14)
        self._header_fill = PatternFill("solid", fgColor="DDE7F5")
        self._center = Alignment(horizontal="center", vertical="center")
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def list_workbook_sheets(self):
        return MAINTENANCE_WORKBOOK_SHEETS

    def list_export_contracts(self):
        return MAINTENANCE_EXPORT_CONTRACTS

    def list_report_contracts(self):
        return MAINTENANCE_REPORT_CONTRACTS

    def export_rollout_workbook_template(
        self,
        output_path: str | Path,
    ):
        output = ensure_output_path(output_path)
        workbook = Workbook()
        index = workbook.active
        index.title = "Workbook Index"
        index["A1"] = "Maintenance Rollout Workbook"
        index["A1"].font = self._title_font
        self._write_header_row(
            index,
            ("Sheet", "Owner", "Operation Key", "Key Field", "Dependencies", "Description"),
            row=3,
        )
        for row_index, contract in enumerate(MAINTENANCE_WORKBOOK_SHEETS, start=4):
            self._write_data_row(
                index,
                row_index,
                (
                    contract.sheet_name,
                    contract.owner_module_code,
                    contract.operation_key or "",
                    contract.key_field,
                    ", ".join(contract.depends_on_sheets),
                    contract.description,
                ),
            )
        self._set_widths(
            index,
            (
                ("A", 24),
                ("B", 24),
                ("C", 34),
                ("D", 20),
                ("E", 34),
                ("F", 72),
            ),
        )

        for contract in MAINTENANCE_WORKBOOK_SHEETS:
            sheet = workbook.create_sheet(contract.sheet_name)
            sheet["A1"] = contract.entity_label
            sheet["A1"].font = self._title_font
            sheet["A2"] = f"Owner: {contract.owner_module_code}"
            sheet["A3"] = contract.description
            if contract.depends_on_sheets:
                sheet["A4"] = f"Depends on: {', '.join(contract.depends_on_sheets)}"
            labels = tuple(field.label for field in contract.field_specs)
            keys = tuple(field.key for field in contract.field_specs)
            self._write_header_row(sheet, labels, row=6)
            self._write_data_row(sheet, 7, keys)
            for column_index, field in enumerate(contract.field_specs, start=1):
                cell = sheet.cell(row=8, column=column_index, value="required" if field.required else "optional")
                cell.alignment = self._center
                cell.border = self._thin_border
            self._set_uniform_widths(sheet, len(contract.field_specs), width=22)

        workbook.save(output)
        return finalize_artifact(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata={
                "artifact_kind": "maintenance_rollout_workbook_template",
                "sheet_count": len(MAINTENANCE_WORKBOOK_SHEETS) + 1,
            },
        )

    def export_runtime_contract_matrix(
        self,
        output_path: str | Path,
    ):
        output = ensure_output_path(output_path)
        with output.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(
                [
                    "contract_group",
                    "key",
                    "label",
                    "owner_module",
                    "format",
                    "dependencies",
                    "description",
                ]
            )
            for contract in MAINTENANCE_WORKBOOK_SHEETS:
                writer.writerow(
                    [
                        "import_sheet",
                        contract.operation_key or contract.sheet_name,
                        contract.sheet_name,
                        contract.owner_module_code,
                        "xlsx",
                        "|".join(contract.depends_on_sheets),
                        contract.description,
                    ]
                )
            for contract in MAINTENANCE_EXPORT_CONTRACTS:
                writer.writerow(
                    [
                        "export_pack",
                        contract.operation_key,
                        contract.label,
                        contract.module_code,
                        contract.artifact_format,
                        "",
                        contract.description,
                    ]
                )
            for contract in MAINTENANCE_REPORT_CONTRACTS:
                writer.writerow(
                    [
                        "report_pack",
                        contract.report_key,
                        contract.label,
                        contract.module_code,
                        "|".join(contract.supported_formats),
                        "",
                        contract.description,
                    ]
                )
        return finalize_artifact(
            output,
            media_type="text/csv",
            metadata={
                "artifact_kind": "maintenance_runtime_contract_matrix",
                "row_count": len(MAINTENANCE_WORKBOOK_SHEETS)
                + len(MAINTENANCE_EXPORT_CONTRACTS)
                + len(MAINTENANCE_REPORT_CONTRACTS),
            },
        )

    def build_contract_catalog_report(self) -> ReportDocument:
        workbook_rows = tuple(
            (
                contract.sheet_name,
                contract.owner_module_code,
                contract.operation_key or "",
                contract.key_field,
                ", ".join(contract.depends_on_sheets),
            )
            for contract in MAINTENANCE_WORKBOOK_SHEETS
        )
        export_rows = tuple(
            (
                contract.operation_key,
                contract.label,
                contract.artifact_format,
            )
            for contract in MAINTENANCE_EXPORT_CONTRACTS
        )
        report_rows = tuple(
            (
                contract.report_key,
                contract.label,
                ", ".join(contract.supported_formats),
            )
            for contract in MAINTENANCE_REPORT_CONTRACTS
        )
        return ReportDocument(
            title="Maintenance Runtime Contract Catalog",
            sections=(
                ReportSection(
                    title="Summary",
                    blocks=(
                        MetricBlock(
                            title="Counts",
                            rows=(
                                MetricRow("Workbook sheets", len(MAINTENANCE_WORKBOOK_SHEETS)),
                                MetricRow("Export packs", len(MAINTENANCE_EXPORT_CONTRACTS)),
                                MetricRow("Report packs", len(MAINTENANCE_REPORT_CONTRACTS)),
                            ),
                        ),
                        TextBlock(
                            title="Purpose",
                            body=(
                                "This catalog freezes maintenance rollout workbook ownership, "
                                "export-pack names, and report-pack keys before live maintenance "
                                "services are implemented."
                            ),
                        ),
                    ),
                ),
                ReportSection(
                    title="Workbook",
                    blocks=(
                        TableBlock(
                            title="Workbook Sheets",
                            columns=("Sheet", "Owner", "Operation Key", "Key Field", "Dependencies"),
                            rows=workbook_rows,
                        ),
                    ),
                ),
                ReportSection(
                    title="Exports",
                    blocks=(
                        TableBlock(
                            title="Export Packs",
                            columns=("Operation Key", "Label", "Format"),
                            rows=export_rows,
                        ),
                    ),
                ),
                ReportSection(
                    title="Reports",
                    blocks=(
                        TableBlock(
                            title="Report Packs",
                            columns=("Report Key", "Label", "Formats"),
                            rows=report_rows,
                        ),
                    ),
                ),
            ),
        )

    def _write_header_row(self, sheet, headers, *, row: int = 1) -> None:
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column_index, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._center
            cell.border = self._thin_border

    def _write_data_row(self, sheet, row_index: int, values) -> None:
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = self._thin_border

    @staticmethod
    def _set_widths(sheet, widths) -> None:
        for column, width in widths:
            sheet.column_dimensions[column].width = width

    @staticmethod
    def _set_uniform_widths(sheet, column_count: int, *, width: float) -> None:
        for index in range(1, column_count + 1):
            letter = chr(64 + index)
            sheet.column_dimensions[letter].width = width


__all__ = ["MaintenanceRuntimeContractCatalogService"]
