from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.modules.inventory_procurement.reporting.models import (
    ProcurementOverviewReport,
    StockStatusReport,
)


class InventoryExcelReportRenderer:
    def __init__(self) -> None:
        self._header_font = Font(bold=True)
        self._title_font = Font(bold=True, size=14)
        self._header_fill = PatternFill("solid", fgColor="DDE7F5")
        self._center = Alignment(horizontal="center")
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def render_stock_status(self, report: StockStatusReport, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"

        self._write_metric_sheet(
            summary,
            title=report.title,
            filters=report.filters,
            summary_metrics=report.summary,
        )

        balances = workbook.create_sheet("Balances")
        headers = (
            "Item Code",
            "Item Name",
            "Storeroom",
            "Storeroom Name",
            "Site",
            "UOM",
            "On Hand",
            "Reserved",
            "Available",
            "On Order",
            "Average Cost",
            "Reorder Required",
            "Last Receipt",
            "Last Issue",
        )
        self._write_header_row(balances, headers)
        for row_index, row in enumerate(report.rows, start=2):
            values = (
                row.item_code,
                row.item_name,
                row.storeroom_code,
                row.storeroom_name,
                row.site_code,
                row.uom,
                float(row.on_hand_qty or 0.0),
                float(row.reserved_qty or 0.0),
                float(row.available_qty or 0.0),
                float(row.on_order_qty or 0.0),
                float(row.average_cost or 0.0),
                "Yes" if row.reorder_required else "No",
                row.last_receipt_at,
                row.last_issue_at,
            )
            self._write_data_row(balances, row_index, values)
        self._set_widths(
            balances,
            (
                ("A", 18),
                ("B", 28),
                ("C", 18),
                ("D", 24),
                ("E", 14),
                ("F", 12),
                ("G", 14),
                ("H", 14),
                ("I", 14),
                ("J", 14),
                ("K", 14),
                ("L", 16),
                ("M", 22),
                ("N", 22),
            ),
        )
        workbook.save(output_path)
        return output_path

    def render_procurement_overview(
        self,
        report: ProcurementOverviewReport,
        output_path: Path,
    ) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"

        self._write_metric_sheet(
            summary,
            title=report.title,
            filters=report.filters,
            summary_metrics=report.summary,
        )

        requisitions = workbook.create_sheet("Requisitions")
        self._write_header_row(
            requisitions,
            (
                "Requisition",
                "Status",
                "Site",
                "Storeroom",
                "Requester",
                "Needed By",
                "Priority",
                "Lines",
                "Requested Qty",
                "Sourced Qty",
                "Purpose",
            ),
        )
        for row_index, row in enumerate(report.requisitions, start=2):
            self._write_data_row(
                requisitions,
                row_index,
                (
                    row.requisition_number,
                    row.status,
                    row.site_code,
                    row.storeroom_code,
                    row.requester_username,
                    row.needed_by_date,
                    row.priority,
                    int(row.line_count),
                    float(row.requested_qty or 0.0),
                    float(row.sourced_qty or 0.0),
                    row.purpose,
                ),
            )
        self._set_widths(
            requisitions,
            (
                ("A", 20),
                ("B", 18),
                ("C", 14),
                ("D", 18),
                ("E", 18),
                ("F", 14),
                ("G", 12),
                ("H", 10),
                ("I", 14),
                ("J", 14),
                ("K", 28),
            ),
        )

        purchase_orders = workbook.create_sheet("Purchase Orders")
        self._write_header_row(
            purchase_orders,
            (
                "PO Number",
                "Status",
                "Site",
                "Supplier",
                "Order Date",
                "Expected Delivery",
                "Lines",
                "Ordered Qty",
                "Received Qty",
                "Rejected Qty",
                "Open Qty",
                "Currency",
            ),
        )
        for row_index, row in enumerate(report.purchase_orders, start=2):
            self._write_data_row(
                purchase_orders,
                row_index,
                (
                    row.po_number,
                    row.status,
                    row.site_code,
                    row.supplier_code,
                    row.order_date,
                    row.expected_delivery_date,
                    int(row.line_count),
                    float(row.ordered_qty or 0.0),
                    float(row.received_qty or 0.0),
                    float(row.rejected_qty or 0.0),
                    float(row.open_qty or 0.0),
                    row.currency_code,
                ),
            )
        self._set_widths(
            purchase_orders,
            (
                ("A", 20),
                ("B", 18),
                ("C", 14),
                ("D", 18),
                ("E", 14),
                ("F", 18),
                ("G", 10),
                ("H", 14),
                ("I", 14),
                ("J", 14),
                ("K", 14),
                ("L", 12),
            ),
        )

        receipts = workbook.create_sheet("Receipts")
        self._write_header_row(
            receipts,
            (
                "Receipt",
                "PO Number",
                "Status",
                "Site",
                "Supplier",
                "Receipt Date",
                "Lines",
                "Accepted Qty",
                "Rejected Qty",
                "Received By",
            ),
        )
        for row_index, row in enumerate(report.receipts, start=2):
            self._write_data_row(
                receipts,
                row_index,
                (
                    row.receipt_number,
                    row.purchase_order_number,
                    row.status,
                    row.site_code,
                    row.supplier_code,
                    row.receipt_date,
                    int(row.line_count),
                    float(row.accepted_qty or 0.0),
                    float(row.rejected_qty or 0.0),
                    row.received_by_username,
                ),
            )
        self._set_widths(
            receipts,
            (
                ("A", 20),
                ("B", 20),
                ("C", 14),
                ("D", 14),
                ("E", 18),
                ("F", 22),
                ("G", 10),
                ("H", 14),
                ("I", 14),
                ("J", 18),
            ),
        )

        workbook.save(output_path)
        return output_path

    def _write_metric_sheet(self, sheet, *, title, filters, summary_metrics) -> None:
        sheet["A1"] = title
        sheet["A1"].font = self._title_font
        row = 3
        row = self._write_metric_block(sheet, row, "Filters", filters)
        row += 1
        self._write_metric_block(sheet, row, "Summary", summary_metrics)
        self._set_widths(sheet, (("A", 30), ("B", 22)))

    def _write_metric_block(self, sheet, row: int, title: str, metrics) -> int:
        sheet[f"A{row}"] = title
        sheet[f"A{row}"].font = self._header_font
        row += 1
        self._write_header_row(sheet, ("Metric", "Value"), row=row)
        row += 1
        for metric in metrics:
            self._write_data_row(sheet, row, (metric.label, metric.value))
            row += 1
        return row

    def _write_header_row(self, sheet, headers, *, row: int = 1) -> None:
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column_index, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.alignment = self._center
            cell.border = self._thin_border

    def _write_data_row(self, sheet, row_index: int, values) -> None:
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.border = self._thin_border

    @staticmethod
    def _set_widths(sheet, widths) -> None:
        for column, width in widths:
            sheet.column_dimensions[column].width = width


__all__ = ["InventoryExcelReportRenderer"]
